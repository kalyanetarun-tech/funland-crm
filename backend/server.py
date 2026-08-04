"""Funland Adventure Park CRM backend."""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')


import os
import uuid
import secrets
import logging
import base64
from datetime import datetime, timezone, date, timedelta
from typing import List, Optional, Literal

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, status, UploadFile, File, Query
from fastapi.responses import Response, PlainTextResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

# -------------- Setup --------------
mongo_url = "mongodb://localhost:27017"
client = AsyncIOMotorClient(
    mongo_url,
    serverSelectionTimeoutMS=5000,
    connectTimeoutMS=5000,
    socketTimeoutMS=15000,
    maxPoolSize=50,
    retryWrites=True,
)
db = client["funland_db"]

JWT_SECRET = "funland_secret_key_123"
JWT_ALGO = "HS256"
JWT_EXPIRE_HOURS = 24 * 7

app = FastAPI(title="Funland CRM")
api = APIRouter(prefix="/api")

@api.get("/health")
async def health():
    """Lightweight health/warm-up endpoint (no auth). Frontend pings this on app load."""
    return {"ok": True, "ts": datetime.now(timezone.utc).isoformat()}

security = HTTPBearer(auto_error=False)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("funland")

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def new_id():
    return str(uuid.uuid4())

def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_pw(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def make_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    if not creds:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except Exception:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    if user.get("role") != "admin" and not user.get("permissions"):
        user["permissions"] = DEFAULT_EMP_PERMS
    if user.get("role") == "admin":
        user["permissions"] = ALL_PERMS
    return user

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return user

# ---------------- Models ----------------
class LoginIn(BaseModel):
    email: EmailStr
    password: str

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    phone: Optional[str] = ""
    role: Literal["admin", "employee"] = "employee"
    permissions: List[str] = []
    is_marketing_exec: bool = False

class UserUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[Literal["admin", "employee"]] = None
    password: Optional[str] = None
    permissions: Optional[List[str]] = None
    is_marketing_exec: Optional[bool] = None

class GameIn(BaseModel):
    name: str
    category: str = "activity"  # top-level item bucket: food / rooms / activities / miscellaneous / games / other
    price: float
    offer_price: Optional[float] = None
    duration_min: Optional[int] = None
    description: Optional[str] = ""
    active: bool = True
    # GST fields (Indian compliance) — kept internal, hidden from customer-facing receipts
    gst_category: Literal["food", "activity", "goods", "room", "clothing", "merchandise", "other"] = "activity"
    hsn_code: Optional[str] = ""

class PackageSplitLine(BaseModel):
    label: str                       # user-facing name shown on invoice
    category: Literal["food", "activity", "room", "clothing", "merchandise", "other"] = "activity"
    amount: float                    # ₹ portion of the package attributable to this line
    gst_percent: Optional[float] = None  # override auto rate if needed
    hsn_code: Optional[str] = ""

class PackageIn(BaseModel):
    name: str
    type: Literal["birthday", "party", "group", "other"] = "birthday"
    category: Optional[str] = ""  # free-text category eg: "Kids Special", "Corporate", "Weekend"
    price: float
    offer_price: Optional[float] = None
    pax: int = 10
    inclusions: List[str] = []
    description: Optional[str] = ""
    active: bool = True
    # NEW — arbitrary breakup lines with their own GST (activity/food/rooms/clothing/other)
    gst_split: List[PackageSplitLine] = []
    # LEGACY — kept for backward compat; auto-migrated to gst_split at bill time
    food_portion: float = 0
    activity_portion: float = 0
    hsn_food: Optional[str] = "996331"
    hsn_activity: Optional[str] = "999721"

class InquiryIn(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    source: Literal["walk-in", "phone", "instagram", "facebook", "whatsapp", "referral", "other"] = "walk-in"
    interest: Optional[str] = ""
    notes: Optional[str] = ""
    follow_up_date: Optional[str] = None
    status: Literal["new", "contacted", "converted", "lost"] = "new"

class InquiryStatusUpdate(BaseModel):
    status: Literal["new", "contacted", "converted", "lost"]
    notes: Optional[str] = None

class InquiryAssign(BaseModel):
    assigned_to: Optional[str] = None  # user id, null to unassign

class RemarkIn(BaseModel):
    text: str

class BillItem(BaseModel):
    kind: Literal["game", "package", "custom"]
    ref_id: Optional[str] = None
    name: str
    price: float
    qty: int = 1
    gst_percent: float = 0  # per-line GST (e.g. food 5%, activity 18%)
    category: Optional[str] = None  # optional label: "food" / "activity" / "entry"
    hsn_code: Optional[str] = ""    # HSN/SAC — for GST invoice

class BillIn(BaseModel):
    customer_name: str
    customer_phone: str = ""
    customer_email: Optional[str] = ""
    customer_gstin: Optional[str] = ""      # 15-char GSTIN if B2B
    customer_state_code: Optional[str] = ""  # 2-digit GST state code (e.g. "23" for MP). Blank = intra-state
    items: List[BillItem]
    discount: float = 0            # flat rupees discount
    discount_percent: float = 0    # OR percent discount (5-100)
    gst_percent: float = 0         # legacy: applied only if no per-item GST provided
    payment_method: Literal["cash", "upi_qr", "razorpay", "card", "rtgs", "netbanking", "cheque", "other"] = "cash"
    payment_status: Literal["pending", "paid"] = "pending"
    # Digital payment audit trail — REQUIRED when payment_method is not cash and status is paid
    payment_reference: Optional[str] = ""   # UPI RRN / txn id / card auth code / cheque no / RTGS UTR
    payment_at: Optional[str] = ""          # ISO datetime of the actual payment
    checked_by: Optional[str] = ""          # Name of staff who verified receipt
    notes: Optional[str] = ""

class AttendanceCheckIn(BaseModel):
    notes: Optional[str] = ""

class PrebookItem(BaseModel):
    kind: Literal["game", "package"]
    ref_id: str
    name: str
    price: float
    qty: int = 1

class PrebookIn(BaseModel):
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = ""
    booking_date: str  # YYYY-MM-DD
    booking_time: Optional[str] = ""  # e.g. "5:30 PM"
    pax: int = 1
    items: List[PrebookItem]
    notes: Optional[str] = ""
    source: str = "web"  # web / whatsapp / phone etc.

class PrebookStatusUpdate(BaseModel):
    status: Literal["pending", "confirmed", "paid", "cancelled", "arrived"]

class SettingsIn(BaseModel):
    park_name: Optional[str] = None
    gst_rate: Optional[float] = None
    upi_qr_url: Optional[str] = None
    upi_id: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    google_review_url: Optional[str] = None
    google_reviews_shown: Optional[int] = None
    google_rating: Optional[float] = None
    # GST / firm compliance
    firm_name: Optional[str] = None            # legal name printed on tax invoice
    firm_gstin: Optional[str] = None           # 15-char GSTIN
    firm_state_code: Optional[str] = None      # 2-digit GST state code, e.g. "23" for MP
    firm_pan: Optional[str] = None
    firm_fssai: Optional[str] = None           # FSSAI for food service
    invoice_prefix: Optional[str] = None       # e.g. "FL/24-25/"

class CampaignIn(BaseModel):
    title: str
    channel: Literal["instagram", "facebook", "whatsapp", "sms", "email"]
    message: str
    image_url: Optional[str] = ""
    audience: Literal["all_customers", "recent_customers", "inquiries", "custom"] = "all_customers"
    custom_phones: List[str] = []

class SendBillIn(BaseModel):
    channel: Literal["whatsapp", "sms", "email"]

# ---------------- Auth ----------------
@api.post("/auth/login")
async def login(data: LoginIn):
    user = await db.users.find_one({"email": data.email.lower()})
    if not user or not verify_pw(data.password, user["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    token = make_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"],
            "phone": user.get("phone", ""),
            "permissions": user.get("permissions") or (ALL_PERMS if user["role"] == "admin" else DEFAULT_EMP_PERMS),
            "is_marketing_exec": user.get("is_marketing_exec", False),
        },
    }

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

# ---------------- Users (staff) - admin ----------------
@api.get("/users")
async def list_users(_: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    return users

DEFAULT_EMP_PERMS = ["dashboard", "inquiries", "visit", "bills", "customers", "games", "packages", "attendance", "prebookings"]
ALL_PERMS = DEFAULT_EMP_PERMS + ["staff", "marketing", "settings"]

@api.post("/users")
async def create_user(data: UserCreate, _: dict = Depends(require_admin)):
    email = data.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already exists")
    perms = data.permissions if data.permissions else (ALL_PERMS if data.role == "admin" else DEFAULT_EMP_PERMS)
    doc = {
        "id": new_id(),
        "email": email,
        "name": data.name,
        "phone": data.phone or "",
        "role": data.role,
        "permissions": perms,
        "is_marketing_exec": data.is_marketing_exec,
        "password_hash": hash_pw(data.password),
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    doc.pop("password_hash")
    doc.pop("_id", None)
    return doc

@api.patch("/users/{uid}")
async def update_user(uid: str, data: UserUpdate, _: dict = Depends(require_admin)):
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "password" in update:
        update["password_hash"] = hash_pw(update.pop("password"))
    # If role is being changed and permissions were not explicitly provided, reset perms
    if "role" in update and "permissions" not in update:
        update["permissions"] = ALL_PERMS if update["role"] == "admin" else DEFAULT_EMP_PERMS
    if update:
        await db.users.update_one({"id": uid}, {"$set": update})
    return await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})

@api.delete("/users/{uid}")
async def delete_user(uid: str, admin: dict = Depends(require_admin)):
    if uid == admin["id"]:
        raise HTTPException(400, "Cannot delete self")
    await db.users.delete_one({"id": uid})
    return {"ok": True}

# ---------------- Games ----------------
@api.get("/games")
async def list_games(user: dict = Depends(get_current_user)):
    games = await db.games.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return games

@api.post("/games")
async def create_game(data: GameIn, _: dict = Depends(require_admin)):
    doc = {"id": new_id(), **data.model_dump(), "created_at": now_iso()}
    await db.games.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/games/{gid}")
async def update_game(gid: str, data: GameIn, _: dict = Depends(require_admin)):
    await db.games.update_one({"id": gid}, {"$set": data.model_dump()})
    return await db.games.find_one({"id": gid}, {"_id": 0})

@api.delete("/games/{gid}")
async def delete_game(gid: str, _: dict = Depends(require_admin)):
    await db.games.delete_one({"id": gid})
    return {"ok": True}

# ---------------- Packages ----------------
@api.get("/packages")
async def list_packages(user: dict = Depends(get_current_user)):
    return await db.packages.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api.post("/packages")
async def create_package(data: PackageIn, _: dict = Depends(require_admin)):
    doc = {"id": new_id(), **data.model_dump(), "created_at": now_iso()}
    await db.packages.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/packages/{pid}")
async def update_package(pid: str, data: PackageIn, _: dict = Depends(require_admin)):
    await db.packages.update_one({"id": pid}, {"$set": data.model_dump()})
    return await db.packages.find_one({"id": pid}, {"_id": 0})

@api.delete("/packages/{pid}")
async def delete_package(pid: str, _: dict = Depends(require_admin)):
    await db.packages.delete_one({"id": pid})
    return {"ok": True}

# ---------------- Inquiries ----------------
@api.get("/inquiries")
async def list_inquiries(include_archived: bool = False, only_archived: bool = False, user: dict = Depends(get_current_user)):
    """Active inquiries by default. Archived (soft-deleted) are hidden unless include_archived=1 or only_archived=1."""
    q = {}
    if only_archived:
        q["is_deleted"] = True
    elif not include_archived:
        q["$or"] = [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]
    return await db.inquiries.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)

# Webhook endpoint - accepts inquiries from external sources (WhatsApp/Twilio, Meta, Zapier, etc.)
async def _pick_next_marketing_exec() -> Optional[dict]:
    """Fair-share pick — chooses the marketing exec who currently has the FEWEST OPEN inquiries.
    Ties are broken by fewest total assigned all-time (so newer joiners catch up),
    then finally by round-robin counter for stability."""
    execs = await db.users.find({"is_marketing_exec": True}, {"_id": 0, "id": 1, "name": 1}).sort("name", 1).to_list(100)
    if not execs:
        return None
    # Open inquiries per exec (new + contacted)
    open_counts = {e["id"]: 0 for e in execs}
    total_counts = {e["id"]: 0 for e in execs}
    async for row in db.inquiries.aggregate([
        {"$match": {"assigned_to": {"$in": [e["id"] for e in execs]}, "$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]}},
        {"$group": {"_id": {"a": "$assigned_to", "s": "$status"}, "n": {"$sum": 1}}},
    ]):
        rid = row["_id"]["a"]
        st = row["_id"]["s"]
        total_counts[rid] = total_counts.get(rid, 0) + row["n"]
        if st in ("new", "contacted"):
            open_counts[rid] = open_counts.get(rid, 0) + row["n"]
    # Increment RR counter (used as tie-breaker for deterministic distribution)
    counter_doc = await db.counters.find_one_and_update(
        {"id": "rr_marketing"}, {"$inc": {"v": 1}}, upsert=True, return_document=True
    )
    rr_idx = ((counter_doc or {}).get("v", 0) - 1) % len(execs)
    # Sort: least open first, then least total, then round-robin order
    ordered = sorted(execs, key=lambda e: (open_counts.get(e["id"], 0), total_counts.get(e["id"], 0), (execs.index(e) - rr_idx) % len(execs)))
    return ordered[0]

async def _create_inquiry_from_normalized(name: str, phone: str, email: str, source: str, message: str, meta: dict = None):
    """Shared helper: create + assign inquiry from normalized fields."""
    src_map = {"whatsapp": "whatsapp", "instagram": "instagram", "facebook": "facebook", "sms": "phone", "phone": "phone", "call": "phone", "web": "other", "twilio_sms": "phone", "twilio_wa": "whatsapp"}
    # Normalise phone: keep digits, strip +91
    if phone:
        import re as _re
        digits = _re.sub(r"\D", "", str(phone))
        if len(digits) > 10 and digits.startswith("91"):
            digits = digits[-10:]
        phone = digits[-10:] if len(digits) >= 10 else digits
    if not name and phone:
        name = "Unknown"
    if not name and not phone:
        return None
    assignee = await _pick_next_marketing_exec()
    doc = {
        "id": new_id(),
        "name": name or "Unknown",
        "phone": phone or "",
        "email": email or "",
        "source": src_map.get(source, "other"),
        "interest": (meta or {}).get("interest", ""),
        "notes": message or "",
        "status": "new",
        "remarks": [],
        "assigned_to": assignee["id"] if assignee else None,
        "assigned_to_name": assignee["name"] if assignee else None,
        "created_by": "webhook",
        "created_by_name": f"{source} webhook",
        "created_at": now_iso(),
        "channel_raw": meta or {},
    }
    await db.inquiries.insert_one(doc)
    doc.pop("_id", None)
    logger.info(f"Webhook inquiry from {source}: {doc['name']} ({doc['phone']}) → {assignee['name'] if assignee else 'unassigned'}")
    return doc

def _verify_webhook_secret(settings: dict, provided: Optional[str]) -> bool:
    expected = (settings or {}).get("inquiry_webhook_secret") or ""
    if not expected:
        return True  # never generated — allow through
    return provided == expected

# --- Meta WhatsApp / Instagram / Messenger Cloud API (defined BEFORE the parameterised /{source} route) ---
@api.get("/inquiries/webhook/meta")
async def meta_webhook_verify(hub_mode: str = Query(None, alias="hub.mode"), hub_verify_token: str = Query(None, alias="hub.verify_token"), hub_challenge: str = Query(None, alias="hub.challenge")):
    """Meta requires GET verification when you register the webhook URL in the app dashboard."""
    settings = await _get_settings()
    if hub_mode == "subscribe" and hub_verify_token == (settings.get("meta_verify_token") or ""):
        return PlainTextResponse(hub_challenge or "ok")
    raise HTTPException(403, "verify_token mismatch")

@api.post("/inquiries/webhook/meta")
async def meta_webhook_inbound(request: Request):
    """Meta Cloud API delivers to this endpoint for WhatsApp / Instagram / FB Messenger.
    We inspect entry[*].changes[*].value / entry[*].messaging[*] and create an inquiry per message."""
    try: payload = await request.json()
    except Exception: payload = {}
    created = []
    entries = payload.get("entry", []) if isinstance(payload, dict) else []
    for entry in entries:
        # WhatsApp Cloud API shape
        for ch in entry.get("changes", []) or []:
            v = ch.get("value") or {}
            contacts = { c.get("wa_id"): (c.get("profile") or {}).get("name") for c in (v.get("contacts") or []) }
            for msg in v.get("messages") or []:
                wa_id = msg.get("from")
                name = contacts.get(wa_id) or ""
                text = ""
                if msg.get("type") == "text":
                    text = (msg.get("text") or {}).get("body", "")
                elif msg.get("type") == "interactive":
                    ir = msg.get("interactive") or {}
                    text = (ir.get("button_reply") or ir.get("list_reply") or {}).get("title", "")
                else:
                    text = f"[{msg.get('type')} message]"
                doc = await _create_inquiry_from_normalized(name, wa_id, "", "whatsapp", text, meta={"raw": msg})
                if doc: created.append(doc["id"])
        # Instagram / Messenger shape
        for m in entry.get("messaging", []) or []:
            sender = (m.get("sender") or {}).get("id") or ""
            text = ((m.get("message") or {}).get("text")) or ""
            src = "instagram"  # FB pages under Meta API deliver to same endpoint; caller can distinguish via entry.id
            doc = await _create_inquiry_from_normalized(f"IG/FB user {sender[-6:]}" if sender else "", "", "", src, text, meta={"raw": m, "sender_id": sender})
            if doc: created.append(doc["id"])
    return {"ok": True, "created": created}


@api.post("/inquiries/webhook/{source}")
async def inquiry_webhook(source: str, request: Request, secret: Optional[str] = Query(None), x_webhook_secret: Optional[str] = None):
    """Public inbound webhook - accepts many payload shapes.

    Auth: pass ?secret=<inquiry_webhook_secret> OR header X-Webhook-Secret.

    Supported payloads (auto-detected):
    - Generic JSON: {name?, phone?, email?, message?/notes?}
    - Twilio SMS/WhatsApp (form-urlencoded): From=..., Body=..., ProfileName?
    - Android SMS Forwarder v3: {from, text, sentStamp, ...}
    - Zapier flat: any of the above keys
    """
    settings = await _get_settings()
    provided = secret or request.headers.get("x-webhook-secret") or request.headers.get("X-Webhook-Secret")
    if not _verify_webhook_secret(settings, provided):
        raise HTTPException(401, "Invalid webhook secret")

    ctype = (request.headers.get("content-type") or "").lower()
    body_payload = {}
    if "application/x-www-form-urlencoded" in ctype or "multipart/form-data" in ctype:
        form = await request.form()
        body_payload = dict(form)
    else:
        try: body_payload = await request.json()
        except Exception: body_payload = {}

    # Normalise across shapes
    name = body_payload.get("name") or body_payload.get("ProfileName") or body_payload.get("profile_name") or body_payload.get("sender_name") or body_payload.get("from_name")
    phone = body_payload.get("phone") or body_payload.get("from") or body_payload.get("From") or body_payload.get("wa_id") or body_payload.get("mobile")
    email = body_payload.get("email") or body_payload.get("Email")
    message = body_payload.get("message") or body_payload.get("Body") or body_payload.get("text") or body_payload.get("notes") or body_payload.get("msg")

    # Twilio WhatsApp: From="whatsapp:+9198..." → normalise
    if phone and isinstance(phone, str) and phone.lower().startswith("whatsapp:"):
        phone = phone.split(":", 1)[1]
        source = "whatsapp"
    # Twilio SMS: From="+9198..."
    if source == "twilio":
        source = "whatsapp" if body_payload.get("MessagingServiceSid", "").lower().startswith("mg") or "whatsapp" in str(body_payload.get("MessageStatus", "")) else "sms"

    doc = await _create_inquiry_from_normalized(name, phone, email, source, message, meta={"raw": body_payload})
    if not doc:
        return {"ok": False, "reason": "no_name_or_phone"}

    # Twilio expects a TwiML response (optional; empty is fine)
    if "twilio" in source or "MessageSid" in body_payload:
        return PlainTextResponse("<Response></Response>", media_type="application/xml")
    return {"ok": True, "id": doc["id"], "assigned_to": doc.get("assigned_to_name")}


@api.post("/inquiries")
async def create_inquiry(data: InquiryIn, user: dict = Depends(get_current_user)):
    assignee = await _pick_next_marketing_exec()
    doc = {
        "id": new_id(),
        **data.model_dump(),
        "remarks": [],
        "assigned_to": assignee["id"] if assignee else None,
        "assigned_to_name": assignee["name"] if assignee else None,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
    }
    await db.inquiries.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/inquiries/{iid}/status")
async def update_inquiry_status(iid: str, data: InquiryStatusUpdate, user: dict = Depends(get_current_user)):
    """Both roles can update status (mark as contacted/converted)."""
    update = {"status": data.status}
    if data.notes is not None:
        update["notes"] = data.notes
    await db.inquiries.update_one({"id": iid}, {"$set": update})
    return await db.inquiries.find_one({"id": iid}, {"_id": 0})

@api.patch("/inquiries/{iid}/assign")
async def assign_inquiry(iid: str, data: InquiryAssign, _: dict = Depends(require_admin)):
    """Admin can reassign inquiry to a specific executive (or unassign with null)."""
    upd = {"assigned_to": None, "assigned_to_name": None}
    if data.assigned_to:
        target = await db.users.find_one({"id": data.assigned_to}, {"_id": 0, "id": 1, "name": 1})
        if not target:
            raise HTTPException(404, "User not found")
        upd = {"assigned_to": target["id"], "assigned_to_name": target["name"]}
    await db.inquiries.update_one({"id": iid}, {"$set": upd})
    return await db.inquiries.find_one({"id": iid}, {"_id": 0})

@api.post("/inquiries/{iid}/remarks")
async def add_remark(iid: str, data: RemarkIn, user: dict = Depends(get_current_user)):
    """Any staff can add a remark (why not converted, follow-up etc.)."""
    if not data.text.strip():
        raise HTTPException(400, "Remark text required")
    remark = {"text": data.text.strip(), "by": user["name"], "by_id": user["id"], "at": now_iso()}
    await db.inquiries.update_one({"id": iid}, {"$push": {"remarks": remark}})
    return await db.inquiries.find_one({"id": iid}, {"_id": 0})

@api.delete("/inquiries/{iid}")
async def soft_delete_inquiry(iid: str, user: dict = Depends(get_current_user)):
    """SOFT delete — moves inquiry to archive, doesn't actually erase.
    Restore via POST /api/inquiries/{iid}/restore. Admin can force-erase via ?permanent=1 (not exposed on this endpoint)."""
    doc = await db.inquiries.find_one({"id": iid}, {"_id": 0, "id": 1})
    if not doc:
        raise HTTPException(404, "Inquiry not found")
    await db.inquiries.update_one({"id": iid}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": user["id"], "deleted_by_name": user["name"]}})
    return {"ok": True, "id": iid, "archived": True}

@api.post("/inquiries/{iid}/restore")
async def restore_inquiry(iid: str, user: dict = Depends(get_current_user)):
    """Un-archive a previously soft-deleted inquiry."""
    r = await db.inquiries.update_one({"id": iid}, {"$set": {"is_deleted": False}, "$unset": {"deleted_at": "", "deleted_by": "", "deleted_by_name": ""}})
    if r.matched_count == 0:
        raise HTTPException(404, "Inquiry not found")
    return await db.inquiries.find_one({"id": iid}, {"_id": 0})

# --- Inquiries Excel import/export ---
INQUIRY_XLSX_COLUMNS = [
    "Name", "Phone", "Email", "Source", "Interest", "Notes",
    "Follow Up Date", "Status", "Assigned To", "Created At",
]

@api.get("/inquiries/export.xlsx")
async def export_inquiries_xlsx(user: dict = Depends(get_current_user)):
    """Download all inquiries as an .xlsx file. Same columns are accepted by the import endpoint."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    rows = await db.inquiries.find({"$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]}, {"_id": 0}).sort("created_at", -1).to_list(20000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    ws.append(INQUIRY_XLSX_COLUMNS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF7A00")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
    for r in rows:
        ws.append([
            r.get("name", ""),
            r.get("phone", ""),
            r.get("email", ""),
            r.get("source", ""),
            r.get("interest", ""),
            r.get("notes", ""),
            r.get("follow_up_date", "") or "",
            r.get("status", ""),
            r.get("assigned_to_name", "") or "",
            (r.get("created_at", "") or "")[:19].replace("T", " "),
        ])
    # Auto width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    buf = BytesIO()
    wb.save(buf)
    filename = f"inquiries_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@api.get("/inquiries/template.xlsx")
async def download_inquiry_template(user: dict = Depends(get_current_user)):
    """Blank template with headers + 2 example rows so users know the format to upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    ws.append(INQUIRY_XLSX_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="FF7A00")
    ws.append(["Rahul Sharma", "9876543210", "rahul@example.com", "whatsapp", "Birthday package", "Sat evening slot", "2026-08-10", "new", "", ""])
    ws.append(["Priya Patel", "9998887777", "", "instagram", "Corporate group", "20 pax", "", "contacted", "", ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inquiries_template.xlsx"'},
    )

@api.post("/inquiries/import")
async def import_inquiries_xlsx(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Bulk import inquiries from ANY .xlsx file — very lenient.
    We auto-detect Name and Phone columns from common aliases (English + Hindi).
    If header is missing or weird, we scan every row and pluck the first phone-like
    cell + the first text cell for name. Phone with no name → 'Unknown'."""
    from openpyxl import load_workbook
    from io import BytesIO
    import re

    if not (file.filename or "").lower().endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(400, "Please upload an .xlsx file")
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot read spreadsheet: {e}")

    valid_sources = {"walk-in", "phone", "instagram", "facebook", "whatsapp", "referral", "other"}
    valid_status = {"new", "contacted", "converted", "lost"}

    NAME_ALIASES   = {"name", "customer", "customer name", "cust name", "full name", "guest name", "guest", "party", "client", "person", "नाम", "ग्राहक", "customer_name"}
    PHONE_ALIASES  = {"phone", "mobile", "mob", "mobile no", "mobile number", "contact", "contact no", "phone no", "number", "no", "no.", "cell", "whatsapp", "wa", "फोन", "मोबाइल", "नंबर", "phone_number", "mobile_number"}
    EMAIL_ALIASES  = {"email", "e-mail", "mail", "email id", "email_id"}
    SOURCE_ALIASES = {"source", "channel", "from", "via", "platform"}
    INTEREST_ALIASES = {"interest", "package", "requirement", "for", "purpose", "event"}
    NOTES_ALIASES  = {"notes", "note", "remark", "remarks", "comment", "comments", "message", "msg", "detail", "description"}
    FOLLOWUP_ALIASES = {"follow up", "follow-up", "followup", "follow up date", "next follow up", "date", "follow_up_date"}
    STATUS_ALIASES = {"status", "stage", "state"}
    ASSIGNED_ALIASES = {"assigned to", "assignee", "assigned", "owner", "executive", "sales", "staff"}

    def norm(v):
        return str(v).strip().lower() if v is not None else ""

    def phone_from(cell) -> str:
        """Return normalised phone number if the cell looks like one, else empty."""
        if cell is None:
            return ""
        s = str(cell).strip()
        if not s:
            return ""
        # Excel may hand us floats for numeric cells (e.g. 9.876543E9)
        if isinstance(cell, float) and cell.is_integer():
            s = str(int(cell))
        # keep only digits
        digits = re.sub(r"\D", "", s)
        if 10 <= len(digits) <= 14:
            # Strip country prefix
            if len(digits) > 10 and digits.startswith("91"):
                digits = digits[-10:]
            if len(digits) == 10 and digits[0] in "6789":
                return digits
        return ""

    def looks_like_name(s: str) -> bool:
        if not s: return False
        s = s.strip()
        if len(s) < 2 or len(s) > 80: return False
        # reject pure numeric / phone-looking / status / email / date
        if re.fullmatch(r"[\d\s\-\+\(\)]+", s): return False
        if "@" in s: return False
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s): return False
        low = s.lower()
        if low in valid_status or low in valid_sources: return False
        return True

    total_inserted = 0
    total_skipped = 0
    errors: List[str] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # -------- Detect header row (may not be row 0) --------
        header_idx = -1
        idx_map = {"name": -1, "phone": -1, "email": -1, "source": -1, "interest": -1, "notes": -1, "follow_up_date": -1, "status": -1, "assigned_to": -1}

        for hi, row in enumerate(rows[:5]):  # search first 5 rows
            cells = [norm(c) for c in row]
            hit_name = any(c in NAME_ALIASES for c in cells)
            hit_phone = any(c in PHONE_ALIASES for c in cells)
            if hit_name or hit_phone:
                header_idx = hi
                for ci, c in enumerate(cells):
                    if c in NAME_ALIASES and idx_map["name"] < 0: idx_map["name"] = ci
                    elif c in PHONE_ALIASES and idx_map["phone"] < 0: idx_map["phone"] = ci
                    elif c in EMAIL_ALIASES and idx_map["email"] < 0: idx_map["email"] = ci
                    elif c in SOURCE_ALIASES and idx_map["source"] < 0: idx_map["source"] = ci
                    elif c in INTEREST_ALIASES and idx_map["interest"] < 0: idx_map["interest"] = ci
                    elif c in NOTES_ALIASES and idx_map["notes"] < 0: idx_map["notes"] = ci
                    elif c in FOLLOWUP_ALIASES and idx_map["follow_up_date"] < 0: idx_map["follow_up_date"] = ci
                    elif c in STATUS_ALIASES and idx_map["status"] < 0: idx_map["status"] = ci
                    elif c in ASSIGNED_ALIASES and idx_map["assigned_to"] < 0: idx_map["assigned_to"] = ci
                break

        data_rows = rows[header_idx + 1:] if header_idx >= 0 else rows

        for row_no, row in enumerate(data_rows, start=(header_idx + 2 if header_idx >= 0 else 1)):
            try:
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue

                # ----- Extract name & phone with fallback scanning -----
                name = ""
                phone = ""

                if idx_map["name"] >= 0 and idx_map["name"] < len(row):
                    v = row[idx_map["name"]]
                    if v is not None and str(v).strip():
                        name = str(v).strip()

                if idx_map["phone"] >= 0 and idx_map["phone"] < len(row):
                    phone = phone_from(row[idx_map["phone"]])

                # Fallback: scan every cell for a phone-shaped value
                if not phone:
                    for c in row:
                        p = phone_from(c)
                        if p:
                            phone = p
                            break

                # Fallback: pick the first name-like text cell (not the phone cell)
                if not name:
                    for ci, c in enumerate(row):
                        if c is None: continue
                        s = str(c).strip()
                        if not s: continue
                        if s == phone or phone_from(c): continue
                        if looks_like_name(s):
                            name = s
                            break

                if not phone and not name:
                    total_skipped += 1
                    continue

                if not name:
                    name = "Unknown"

                # ----- Optional fields -----
                def cell(k):
                    i = idx_map.get(k, -1)
                    if i < 0 or i >= len(row): return ""
                    v = row[i]
                    return str(v).strip() if v is not None else ""

                email = cell("email")
                if email and "@" not in email:
                    # scan any other cell for a valid-ish email
                    for c in row:
                        if c and "@" in str(c):
                            email = str(c).strip(); break

                source = cell("source").lower().replace(" ", "-") or "other"
                if source not in valid_sources:
                    # allow "insta" → instagram etc
                    if source.startswith("insta"): source = "instagram"
                    elif source.startswith("face") or source == "fb": source = "facebook"
                    elif source.startswith("what") or source == "wa": source = "whatsapp"
                    elif source in {"walkin", "walk_in"}: source = "walk-in"
                    else: source = "other"

                status_val = cell("status").lower() or "new"
                if status_val not in valid_status:
                    status_val = "new"

                assignee_name = cell("assigned_to")
                assignee_doc = None
                if assignee_name:
                    assignee_doc = await db.users.find_one({"name": assignee_name}, {"_id": 0, "id": 1, "name": 1})
                if not assignee_doc:
                    assignee_doc = await _pick_next_marketing_exec()

                doc = {
                    "id": new_id(),
                    "name": name,
                    "phone": phone,
                    "email": email,
                    "source": source,
                    "interest": cell("interest"),
                    "notes": cell("notes"),
                    "follow_up_date": cell("follow_up_date") or None,
                    "status": status_val,
                    "remarks": [],
                    "assigned_to": (assignee_doc or {}).get("id"),
                    "assigned_to_name": (assignee_doc or {}).get("name"),
                    "created_by": user["id"],
                    "created_by_name": user["name"] + " (import)",
                    "created_at": now_iso(),
                }
                await db.inquiries.insert_one(doc)
                total_inserted += 1
            except Exception as e:
                errors.append(f"{ws.title} row {row_no}: {e}")

    return {"ok": True, "inserted": total_inserted, "skipped": total_skipped, "errors": errors[:20]}


# ---------------- Expenses ----------------
class ExpenseIn(BaseModel):
    date: str                          # ISO YYYY-MM-DD
    category: str = "other"           # rent, salary, utility, food, maintenance, marketing, other
    description: str = ""
    amount: float
    payment_method: str = "cash"      # cash / upi / bank / cheque / other
    payment_reference: Optional[str] = ""
    vendor: Optional[str] = ""
    bill_url: Optional[str] = ""       # attached vendor bill image URL

@api.get("/expenses")
async def list_expenses(user: dict = Depends(get_current_user)):
    return await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(5000)

@api.post("/expenses")
async def create_expense(data: ExpenseIn, user: dict = Depends(get_current_user)):
    doc = {"id": new_id(), **data.model_dump(), "created_by": user["id"], "created_by_name": user["name"], "created_at": now_iso()}
    await db.expenses.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/expenses/{eid}")
async def update_expense(eid: str, data: ExpenseIn, user: dict = Depends(get_current_user)):
    await db.expenses.update_one({"id": eid}, {"$set": data.model_dump()})
    return await db.expenses.find_one({"id": eid}, {"_id": 0})

@api.delete("/expenses/{eid}")
async def delete_expense(eid: str, user: dict = Depends(get_current_user)):
    r = await db.expenses.delete_one({"id": eid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"ok": True}


# ---------------- Business Reports (Sales / GST-3B / Payment / Expense) ----------------
async def _fetch_bills_in_range(f_iso: str, t_iso: str):
    """Bills with created_at in [f_iso, t_iso). Only paid bills contribute to revenue/GST."""
    return await db.bills.find({"created_at": {"$gte": f_iso, "$lt": t_iso}}, {"_id": 0}).to_list(50000)

@api.get("/reports/sales")
async def sales_report(from_date: Optional[str] = Query(None, alias="from"), to_date: Optional[str] = Query(None, alias="to"), preset: Optional[str] = None, user: dict = Depends(get_current_user)):
    f_iso, t_iso, label = _parse_range_yyyymmdd(from_date, to_date, preset)
    bills = await _fetch_bills_in_range(f_iso, t_iso)
    total_revenue = 0.0
    total_bills = 0
    total_paid = 0
    total_pending = 0
    category_totals = {}      # by item category (food/activity/room/…): {taxable, tax, gross}
    daily = {}                # date → revenue
    top_items = {}
    for b in bills:
        total_bills += 1
        if b.get("payment_status") == "paid":
            amt = float(b.get("total") or 0)
            total_revenue += amt
            total_paid += 1
            d = (b.get("created_at") or "")[:10]
            if d: daily[d] = daily.get(d, 0) + amt
        else:
            total_pending += 1
        for it in b.get("items", []):
            cat = (it.get("category") or "other").lower()
            gross = float(it.get("price") or 0) * int(it.get("qty") or 0)
            rate = float(it.get("gst_percent") or 0)
            taxable = gross / (1 + rate/100) if rate else gross
            tax = gross - taxable
            slot = category_totals.setdefault(cat, {"taxable": 0.0, "tax": 0.0, "gross": 0.0, "count": 0})
            slot["taxable"] += taxable
            slot["tax"] += tax
            slot["gross"] += gross
            slot["count"] += int(it.get("qty") or 0)
            top_items[it.get("name", "?")] = top_items.get(it.get("name", "?"), 0) + int(it.get("qty") or 0)
    daily_list = sorted([{"date": d, "revenue": round(v, 2)} for d, v in daily.items()], key=lambda x: x["date"])
    top = sorted([{"name": k, "qty": v} for k, v in top_items.items()], key=lambda x: -x["qty"])[:10]
    for k, v in category_totals.items():
        v["taxable"] = round(v["taxable"], 2)
        v["tax"] = round(v["tax"], 2)
        v["gross"] = round(v["gross"], 2)
    return {
        "from": f_iso, "to": t_iso, "label": label,
        "total_revenue": round(total_revenue, 2),
        "total_bills": total_bills,
        "paid_bills": total_paid,
        "pending_bills": total_pending,
        "avg_bill_value": round(total_revenue / total_paid, 2) if total_paid else 0,
        "category_totals": category_totals,
        "daily": daily_list,
        "top_items": top,
    }

@api.get("/reports/gstr3b")
async def gstr3b_report(from_date: Optional[str] = Query(None, alias="from"), to_date: Optional[str] = Query(None, alias="to"), preset: Optional[str] = None, user: dict = Depends(get_current_user)):
    """GSTR-3B style outward supplies summary: per-rate taxable, CGST, SGST, IGST."""
    f_iso, t_iso, label = _parse_range_yyyymmdd(from_date, to_date, preset)
    bills = await _fetch_bills_in_range(f_iso, t_iso)
    by_rate = {}
    total_taxable = 0.0
    total_cgst = 0.0
    total_sgst = 0.0
    total_igst = 0.0
    invoices = 0
    for b in bills:
        if b.get("payment_status") != "paid":
            continue
        invoices += 1
        for br in (b.get("gst_breakup") or []):
            rate = float(br.get("rate") or 0)
            slot = by_rate.setdefault(rate, {"rate": rate, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0})
            slot["taxable"] += float(br.get("taxable") or 0)
            slot["cgst"] += float(br.get("cgst") or 0)
            slot["sgst"] += float(br.get("sgst") or 0)
            slot["igst"] += float(br.get("igst") or 0)
            total_taxable += float(br.get("taxable") or 0)
            total_cgst += float(br.get("cgst") or 0)
            total_sgst += float(br.get("sgst") or 0)
            total_igst += float(br.get("igst") or 0)
    breakup = sorted([{**v, "taxable": round(v["taxable"], 2), "cgst": round(v["cgst"], 2), "sgst": round(v["sgst"], 2), "igst": round(v["igst"], 2)} for v in by_rate.values()], key=lambda x: x["rate"])
    return {
        "from": f_iso, "to": t_iso, "label": label,
        "invoice_count": invoices,
        "total_taxable": round(total_taxable, 2),
        "total_cgst": round(total_cgst, 2),
        "total_sgst": round(total_sgst, 2),
        "total_igst": round(total_igst, 2),
        "total_tax": round(total_cgst + total_sgst + total_igst, 2),
        "rate_wise": breakup,
    }

@api.get("/reports/payment-mode")
async def payment_mode_report(from_date: Optional[str] = Query(None, alias="from"), to_date: Optional[str] = Query(None, alias="to"), preset: Optional[str] = None, user: dict = Depends(get_current_user)):
    f_iso, t_iso, label = _parse_range_yyyymmdd(from_date, to_date, preset)
    bills = await _fetch_bills_in_range(f_iso, t_iso)
    by_mode = {}
    total_paid = 0.0
    total_pending = 0.0
    for b in bills:
        method = b.get("payment_method") or "cash"
        amt = float(b.get("total") or 0)
        slot = by_mode.setdefault(method, {"method": method, "paid_amount": 0.0, "paid_count": 0, "pending_amount": 0.0, "pending_count": 0})
        if b.get("payment_status") == "paid":
            slot["paid_amount"] += amt
            slot["paid_count"] += 1
            total_paid += amt
        else:
            slot["pending_amount"] += amt
            slot["pending_count"] += 1
            total_pending += amt
    modes = sorted([{**v, "paid_amount": round(v["paid_amount"], 2), "pending_amount": round(v["pending_amount"], 2)} for v in by_mode.values()], key=lambda x: -x["paid_amount"])
    return {
        "from": f_iso, "to": t_iso, "label": label,
        "total_paid": round(total_paid, 2),
        "total_pending": round(total_pending, 2),
        "modes": modes,
    }

@api.get("/reports/expenses")
async def expense_report(from_date: Optional[str] = Query(None, alias="from"), to_date: Optional[str] = Query(None, alias="to"), preset: Optional[str] = None, user: dict = Depends(get_current_user)):
    f_iso, t_iso, label = _parse_range_yyyymmdd(from_date, to_date, preset)
    rows = await db.expenses.find({"date": {"$gte": f_iso[:10], "$lt": t_iso[:10]}}, {"_id": 0}).sort("date", -1).to_list(20000)
    by_cat = {}
    by_month = {}
    total = 0.0
    for r in rows:
        cat = (r.get("category") or "other").lower()
        amt = float(r.get("amount") or 0)
        total += amt
        by_cat[cat] = by_cat.get(cat, 0) + amt
        m = r.get("date", "")[:7]
        by_month[m] = by_month.get(m, 0) + amt
    cats = sorted([{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()], key=lambda x: -x["amount"])
    months = sorted([{"month": k, "amount": round(v, 2)} for k, v in by_month.items()], key=lambda x: x["month"])
    return {
        "from": f_iso, "to": t_iso, "label": label,
        "total": round(total, 2),
        "count": len(rows),
        "by_category": cats,
        "by_month": months,
        "expenses": rows,
    }

@api.get("/reports/business.xlsx")
async def business_xlsx(from_date: Optional[str] = Query(None, alias="from"), to_date: Optional[str] = Query(None, alias="to"), preset: Optional[str] = None, user: dict = Depends(get_current_user)):
    """One workbook, four sheets: Sales, GSTR-3B, Payment Modes, Expenses."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    sales   = await sales_report(from_date=from_date, to_date=to_date, preset=preset, user=user)
    gst     = await gstr3b_report(from_date=from_date, to_date=to_date, preset=preset, user=user)
    pay     = await payment_mode_report(from_date=from_date, to_date=to_date, preset=preset, user=user)
    exp     = await expense_report(from_date=from_date, to_date=to_date, preset=preset, user=user)

    wb = Workbook()
    settings = await _get_settings()
    park = settings.get("firm_name") or settings.get("park_name") or "Funland"

    def header(ws, cols):
        ws.append(cols)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="FF7A00")

    # -- Sales --
    ws = wb.active; ws.title = "Sales"
    ws.append([f"{park} — Sales Report", sales["label"]]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Range: {sales['from']} → {sales['to']}"]); ws.append([])
    ws.append(["Total revenue", sales["total_revenue"]])
    ws.append(["Bills paid", sales["paid_bills"]])
    ws.append(["Bills pending", sales["pending_bills"]])
    ws.append(["Avg bill value", sales["avg_bill_value"]]); ws.append([])
    header(ws, ["Category", "Qty", "Taxable", "Tax", "Gross"])
    for cat, v in sales["category_totals"].items():
        ws.append([cat, v["count"], v["taxable"], v["tax"], v["gross"]])
    ws.append([]); header(ws, ["Date", "Revenue"])
    for d in sales["daily"]:
        ws.append([d["date"], d["revenue"]])
    ws.append([]); header(ws, ["Top item", "Qty sold"])
    for it in sales["top_items"]:
        ws.append([it["name"], it["qty"]])

    # -- GSTR-3B --
    ws2 = wb.create_sheet(title="GSTR-3B")
    ws2.append([f"{park} — GSTR-3B", gst["label"]]); ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([f"Range: {gst['from']} → {gst['to']}"]); ws2.append([])
    ws2.append(["Invoices", gst["invoice_count"]])
    ws2.append(["Total taxable", gst["total_taxable"]])
    ws2.append(["Total CGST", gst["total_cgst"]])
    ws2.append(["Total SGST", gst["total_sgst"]])
    ws2.append(["Total IGST", gst["total_igst"]])
    ws2.append(["Total tax", gst["total_tax"]]); ws2.append([])
    header(ws2, ["Rate %", "Taxable value", "CGST", "SGST", "IGST"])
    for r in gst["rate_wise"]:
        ws2.append([r["rate"], r["taxable"], r["cgst"], r["sgst"], r["igst"]])

    # -- Payment modes --
    ws3 = wb.create_sheet(title="Payment Modes")
    ws3.append([f"{park} — Payment Mode Report", pay["label"]]); ws3["A1"].font = Font(bold=True, size=14)
    ws3.append([f"Range: {pay['from']} → {pay['to']}"]); ws3.append([])
    ws3.append(["Total collected (paid)", pay["total_paid"]])
    ws3.append(["Pending amount", pay["total_pending"]]); ws3.append([])
    header(ws3, ["Method", "Paid count", "Paid amount", "Pending count", "Pending amount"])
    for m in pay["modes"]:
        ws3.append([m["method"], m["paid_count"], m["paid_amount"], m["pending_count"], m["pending_amount"]])

    # -- Expenses --
    ws4 = wb.create_sheet(title="Expenses")
    ws4.append([f"{park} — Expense Report", exp["label"]]); ws4["A1"].font = Font(bold=True, size=14)
    ws4.append([f"Range: {exp['from']} → {exp['to']}"]); ws4.append([])
    ws4.append(["Total expenses", exp["total"]])
    ws4.append(["Entries", exp["count"]]); ws4.append([])
    header(ws4, ["Category", "Amount"])
    for c in exp["by_category"]:
        ws4.append([c["category"], c["amount"]])
    ws4.append([]); header(ws4, ["Date", "Category", "Description", "Vendor", "Method", "Ref", "Amount"])
    for e in exp["expenses"]:
        ws4.append([e.get("date",""), e.get("category",""), e.get("description",""), e.get("vendor",""), e.get("payment_method",""), e.get("payment_reference",""), e.get("amount",0)])

    for wsx in wb.worksheets:
        for col in wsx.columns:
            wsx.column_dimensions[col[0].column_letter].width = 18

    buf = BytesIO(); wb.save(buf)
    filename = f"business_report_{sales['from']}_{sales['to'][:10]}.xlsx".replace(" ", "_")
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})



# ---------------- Marketing Reports ----------------
def _parse_range_yyyymmdd(from_date: Optional[str], to_date: Optional[str], preset: Optional[str]):
    """Return (from_iso, to_iso_exclusive, label) for use in Mongo string compare of created_at."""
    today = datetime.now(timezone.utc).date()
    if preset:
        p = preset.lower()
        if p == "today":
            f, t = today, today; label = "Today"
        elif p in ("week", "7d"):
            f, t = today - timedelta(days=6), today; label = "Last 7 days"
        elif p in ("month", "30d"):
            f, t = today - timedelta(days=29), today; label = "Last 30 days"
        elif p in ("year", "365d"):
            f, t = today - timedelta(days=364), today; label = "Last 12 months"
        elif p == "all":
            f, t = date(2020, 1, 1), today; label = "All time"
        else:
            f, t = today - timedelta(days=29), today; label = "Last 30 days"
    else:
        try:
            f = datetime.fromisoformat(from_date).date() if from_date else today - timedelta(days=29)
            t = datetime.fromisoformat(to_date).date() if to_date else today
        except Exception:
            f, t = today - timedelta(days=29), today
        label = f"{f} → {t}"
    return f.isoformat(), (t + timedelta(days=1)).isoformat(), label


@api.get("/marketing/report")
async def marketing_report(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    preset: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Per-executive marketing performance report in the given window.

    Returns each exec's: assigned, new, contacted, converted, lost, conversion_rate,
    remarks_added, avg_response_hours (time to first remark by that exec),
    source_breakdown, and a per-day trend."""
    f_iso, t_iso, label = _parse_range_yyyymmdd(from_date, to_date, preset)

    execs = await db.users.find({"is_marketing_exec": True}, {"_id": 0, "id": 1, "name": 1, "email": 1}).sort("name", 1).to_list(200)
    inquiries = await db.inquiries.find(
        {"created_at": {"$gte": f_iso, "$lt": t_iso}, "$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]},
        {"_id": 0}
    ).to_list(50000)

    per_exec = {e["id"]: {
        **e,
        "assigned": 0, "new": 0, "contacted": 0, "converted": 0, "lost": 0,
        "remarks_added": 0,
        "response_hours_sum": 0.0, "response_samples": 0,
        "source_breakdown": {},
        "day_trend": {},
    } for e in execs}
    unassigned = {"assigned": 0, "new": 0, "contacted": 0, "converted": 0, "lost": 0}
    totals = {"assigned": 0, "new": 0, "contacted": 0, "converted": 0, "lost": 0}

    for inq in inquiries:
        aid = inq.get("assigned_to")
        status_val = (inq.get("status") or "new").lower()
        bucket = per_exec.get(aid) if aid in per_exec else unassigned
        bucket["assigned"] = bucket.get("assigned", 0) + 1
        bucket[status_val] = bucket.get(status_val, 0) + 1
        totals["assigned"] += 1
        totals[status_val] = totals.get(status_val, 0) + 1
        if aid in per_exec:
            e = per_exec[aid]
            src = inq.get("source", "other") or "other"
            e["source_breakdown"][src] = e["source_breakdown"].get(src, 0) + 1
            day = (inq.get("created_at") or "")[:10]
            if day:
                d = e["day_trend"].setdefault(day, {"assigned": 0, "converted": 0})
                d["assigned"] += 1
                if status_val == "converted":
                    d["converted"] += 1
            remarks = inq.get("remarks") or []
            first_action = None
            for r in remarks:
                if r.get("by_id") == aid:
                    e["remarks_added"] += 1
                    if r.get("at") and (first_action is None or r["at"] < first_action):
                        first_action = r["at"]
            if first_action and inq.get("created_at"):
                try:
                    delta = datetime.fromisoformat(first_action.replace("Z", "+00:00")) - datetime.fromisoformat(inq["created_at"].replace("Z", "+00:00"))
                    hours = max(delta.total_seconds() / 3600.0, 0)
                    e["response_hours_sum"] += hours
                    e["response_samples"] += 1
                except Exception:
                    pass

    executives = []
    for e in per_exec.values():
        assigned = e["assigned"]
        converted = e.get("converted", 0)
        conv_rate = round((converted / assigned) * 100, 1) if assigned else 0.0
        avg_resp = round(e["response_hours_sum"] / e["response_samples"], 1) if e["response_samples"] else None
        trend = sorted([{"date": d, **v} for d, v in e["day_trend"].items()], key=lambda x: x["date"])
        executives.append({
            "id": e["id"], "name": e["name"], "email": e.get("email"),
            "assigned": assigned, "new": e.get("new", 0), "contacted": e.get("contacted", 0),
            "converted": converted, "lost": e.get("lost", 0),
            "conversion_rate": conv_rate,
            "remarks_added": e["remarks_added"],
            "avg_response_hours": avg_resp,
            "source_breakdown": e["source_breakdown"],
            "day_trend": trend,
        })
    executives.sort(key=lambda x: (-x["conversion_rate"], -x["assigned"]))

    totals["conversion_rate"] = round((totals["converted"] / totals["assigned"]) * 100, 1) if totals["assigned"] else 0.0
    return {
        "from": f_iso, "to": t_iso, "label": label,
        "executives": executives, "totals": totals, "unassigned": unassigned,
    }


@api.get("/marketing/report.xlsx")
async def marketing_report_xlsx(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    preset: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    report = await marketing_report(from_date=from_date, to_date=to_date, preset=preset, user=user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    settings = await _get_settings()
    park = settings.get("firm_name") or settings.get("park_name") or "Funland"

    ws.append([f"{park} — Marketing Report", report.get("label", "")])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Range: {report['from']} to {report['to']}"])
    ws.append([])

    headers = ["Executive", "Email", "Assigned", "New", "Contacted", "Converted", "Lost", "Conv. Rate %", "Remarks", "Avg Response (hrs)"]
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="FF7A00")
        c.alignment = Alignment(horizontal="center")

    for e in report["executives"]:
        ws.append([
            e["name"], e.get("email", ""),
            e["assigned"], e["new"], e["contacted"], e["converted"], e["lost"],
            e["conversion_rate"], e["remarks_added"],
            e["avg_response_hours"] if e["avg_response_hours"] is not None else "—",
        ])
    t = report["totals"]
    ws.append([])
    ws.append(["TOTAL", "", t["assigned"], t.get("new", 0), t.get("contacted", 0), t.get("converted", 0), t.get("lost", 0), t.get("conversion_rate", 0), "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFE5CC")

    for e in report["executives"]:
        sheet_name = (e["name"] or "Exec")[:28]
        try:
            ws2 = wb.create_sheet(title=sheet_name)
        except Exception:
            ws2 = wb.create_sheet(title=f"Exec-{report['executives'].index(e)}")
        ws2.append([e["name"]])
        ws2["A1"].font = Font(bold=True, size=13)
        ws2.append([])
        ws2.append(["Source", "Count"])
        for c in ws2[ws2.max_row]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0080FF")
        for src, n in sorted(e["source_breakdown"].items(), key=lambda x: -x[1]):
            ws2.append([src, n])
        ws2.append([]); ws2.append([]); ws2.append(["Date", "Assigned", "Converted"])
        for c in ws2[ws2.max_row]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0080FF")
        for d in e["day_trend"]:
            ws2.append([d["date"], d.get("assigned", 0), d.get("converted", 0)])
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 18

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    filename = f"marketing_report_{report['from']}_{report['to'][:10]}.xlsx".replace(" ", "_")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------- Bills / Visits ----------------
# ---------------- GST helpers (Indian compliance) ----------------
GST_RATE_BY_CATEGORY = {
    "food": 5.0,          # restaurant / catering
    "activity": 18.0,     # amusement / rides
    "goods": 18.0,        # generic goods
    "room": 12.0,         # hotel room < 7500/night
    "clothing": 12.0,     # apparel > 1000
    "merchandise": 18.0,  # merchandise
    "other": 18.0,
}
HSN_BY_CATEGORY = {
    "food": "996331",
    "activity": "999721",
    "goods": "999799",
    "room": "996311",
    "clothing": "6109",
    "merchandise": "999799",
    "other": "999799",
}

def _resolve_line_gst(it: dict) -> (float, str, str):
    """Return (rate, hsn, category) for a bill line, filling from category/kind if missing."""
    rate = float(it.get("gst_percent") or 0)
    cat = (it.get("category") or "").lower()
    hsn = it.get("hsn_code") or ""
    if rate <= 0 and cat in GST_RATE_BY_CATEGORY:
        rate = GST_RATE_BY_CATEGORY[cat]
    if not hsn and cat in HSN_BY_CATEGORY:
        hsn = HSN_BY_CATEGORY[cat]
    if rate <= 0 and it.get("kind") == "game":
        # default for games/activities
        rate = 18.0
        cat = cat or "activity"
        hsn = hsn or HSN_BY_CATEGORY["activity"]
    return rate, hsn, cat

def _compute_bill_totals(items, discount, discount_percent, legacy_gst_percent, is_interstate: bool = False):
    """Compute subtotal / discount / GST breakup / total.
    - Each item can carry its own gst_percent (5% food, 18% activity).
    - GST is calculated on the taxable amount (after proportional discount).
    - Breakup groups by rate; splits into CGST/SGST (intra-state) or IGST (inter-state).
    Returns (subtotal, discount_amount, gst_amount, total, gst_breakup)."""
    subtotal = round(sum(i["price"] * i["qty"] for i in items), 2)
    if discount_percent and discount_percent > 0:
        pct = min(max(discount_percent, 0), 100)
        disc_amount = round(subtotal * pct / 100.0, 2)
    else:
        disc_amount = round(min(max(discount, 0), subtotal), 2)
    after_discount = max(subtotal - disc_amount, 0)
    ratio = (after_discount / subtotal) if subtotal > 0 else 0

    has_line_gst = any((i.get("gst_percent") or 0) > 0 for i in items)
    rate_map = {}   # rate -> {taxable, tax}
    gst_amount = 0.0

    if has_line_gst:
        for it in items:
            rate = float(it.get("gst_percent") or 0)
            line_gross = it["price"] * it["qty"]
            line_taxable = round(line_gross * ratio, 2)
            line_tax = round(line_taxable * rate / 100.0, 2)
            if rate > 0:
                slot = rate_map.setdefault(rate, {"taxable": 0.0, "tax": 0.0})
                slot["taxable"] = round(slot["taxable"] + line_taxable, 2)
                slot["tax"] = round(slot["tax"] + line_tax, 2)
                gst_amount += line_tax
        gst_amount = round(gst_amount, 2)
    elif legacy_gst_percent and legacy_gst_percent > 0:
        gst_amount = round(after_discount * (legacy_gst_percent / 100.0), 2)
        rate_map[float(legacy_gst_percent)] = {"taxable": after_discount, "tax": gst_amount}

    breakup = []
    for rate in sorted(rate_map.keys()):
        row = rate_map[rate]
        if is_interstate:
            breakup.append({
                "rate": rate,
                "taxable": row["taxable"],
                "cgst": 0.0, "sgst": 0.0,
                "igst": round(row["tax"], 2),
                "total_tax": round(row["tax"], 2),
            })
        else:
            half = round(row["tax"] / 2.0, 2)
            breakup.append({
                "rate": rate,
                "taxable": row["taxable"],
                "cgst": half,
                "sgst": round(row["tax"] - half, 2),
                "igst": 0.0,
                "total_tax": round(row["tax"], 2),
            })

    total = round(after_discount + gst_amount, 2)
    return subtotal, disc_amount, gst_amount, total, breakup

def _bill_number():
    return "FL-" + datetime.now(timezone.utc).strftime("%y%m%d") + "-" + uuid.uuid4().hex[:5].upper()

@api.get("/bills")
async def list_bills(user: dict = Depends(get_current_user)):
    return await db.bills.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)

@api.get("/bills/{bid}")
async def get_bill(bid: str, user: dict = Depends(get_current_user)):
    b = await db.bills.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Bill not found")
    return b

@api.post("/bills")
async def create_bill(data: BillIn, user: dict = Depends(get_current_user)):
    # Expand package items into food/activity lines with proper GST rates
    raw_items = [i.model_dump() for i in data.items]
    expanded: List[dict] = []
    # Pre-fetch referenced game/package docs for HSN/category enrichment
    pkg_ids = {i.get("ref_id") for i in raw_items if i.get("kind") == "package" and i.get("ref_id")}
    game_ids = {i.get("ref_id") for i in raw_items if i.get("kind") == "game" and i.get("ref_id")}
    pkgs = {p["id"]: p for p in await db.packages.find({"id": {"$in": list(pkg_ids)}}, {"_id": 0}).to_list(500)} if pkg_ids else {}
    games = {g["id"]: g for g in await db.games.find({"id": {"$in": list(game_ids)}}, {"_id": 0}).to_list(500)} if game_ids else {}

    for it in raw_items:
        if it["kind"] == "package" and it.get("ref_id") and it["ref_id"] in pkgs:
            p = pkgs[it["ref_id"]]
            qty = it.get("qty", 1)
            splits = list(p.get("gst_split") or [])
            # Legacy migration: food_portion + activity_portion → 2-line split
            if not splits:
                fp = float(p.get("food_portion") or 0)
                ap = float(p.get("activity_portion") or 0)
                if fp > 0:
                    splits.append({"label": "Food", "category": "food", "amount": fp, "hsn_code": p.get("hsn_food")})
                if ap > 0:
                    splits.append({"label": "Activity", "category": "activity", "amount": ap, "hsn_code": p.get("hsn_activity")})
            if splits:
                for s in splits:
                    cat = (s.get("category") or "activity").lower()
                    rate = float(s.get("gst_percent") or GST_RATE_BY_CATEGORY.get(cat, 18.0))
                    hsn = s.get("hsn_code") or HSN_BY_CATEGORY.get(cat, "999721")
                    expanded.append({
                        "kind": "package", "ref_id": it["ref_id"],
                        "name": f"{it['name']} · {s.get('label') or cat.title()}",
                        "price": round(float(s.get("amount") or 0), 2),
                        "qty": qty,
                        "gst_percent": rate, "category": cat, "hsn_code": hsn,
                    })
                continue
            # Fallback — single-category package (default activity 18%)
            it["gst_percent"] = it.get("gst_percent") or 18.0
            it["category"] = it.get("category") or "activity"
            it["hsn_code"] = it.get("hsn_code") or HSN_BY_CATEGORY["activity"]
            expanded.append(it)
        elif it["kind"] == "game" and it.get("ref_id") and it["ref_id"] in games:
            g = games[it["ref_id"]]
            cat = (g.get("gst_category") or "activity").lower()
            it["gst_percent"] = it.get("gst_percent") or GST_RATE_BY_CATEGORY.get(cat, 18.0)
            it["category"] = it.get("category") or cat
            it["hsn_code"] = it.get("hsn_code") or (g.get("hsn_code") or HSN_BY_CATEGORY.get(cat, "999721"))
            expanded.append(it)
        else:
            # Custom line — fill category-derived defaults
            rate, hsn, cat = _resolve_line_gst(it)
            it["gst_percent"] = rate or it.get("gst_percent") or 0
            it["category"] = cat or it.get("category")
            it["hsn_code"] = hsn or it.get("hsn_code") or ""
            expanded.append(it)

    # Determine intra vs inter-state
    settings = await _get_settings()
    firm_sc = (settings.get("firm_state_code") or "").strip()
    cust_sc = (data.customer_state_code or "").strip()
    is_interstate = bool(firm_sc and cust_sc and firm_sc != cust_sc)

    subtotal, disc_amount, gst_amount, total, gst_breakup = _compute_bill_totals(
        expanded, data.discount, data.discount_percent, data.gst_percent, is_interstate
    )

    # Enforce digital payment audit trail — if paid via non-cash, checked_by + payment_reference are required
    _digital_methods = {"upi_qr", "razorpay", "card", "rtgs", "netbanking", "cheque"}
    if data.payment_status == "paid" and data.payment_method in _digital_methods:
        if not (data.checked_by or "").strip():
            raise HTTPException(400, "checked_by is compulsory for digital payments — batao kisne verify kiya")
        if not (data.payment_reference or "").strip():
            raise HTTPException(400, "payment_reference (txn id / UTR / RRN / cheque no) compulsory hai non-cash payments ke liye")

    doc = {
        "id": new_id(),
        "bill_no": (settings.get("invoice_prefix") or "") + _bill_number(),
        "customer_name": data.customer_name,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email,
        "customer_gstin": (data.customer_gstin or "").strip().upper(),
        "customer_state_code": cust_sc,
        "items": expanded,
        "discount": disc_amount,
        "discount_percent": data.discount_percent,
        "gst_percent": data.gst_percent,
        "gst_amount": gst_amount,
        "gst_breakup": gst_breakup,
        "is_interstate": is_interstate,
        "subtotal": subtotal,
        "total": total,
        "payment_method": data.payment_method,
        "payment_status": data.payment_status,
        "payment_reference": (data.payment_reference or "").strip(),
        "payment_at": (data.payment_at or "").strip() or (now_iso() if data.payment_status == "paid" else ""),
        "checked_by": (data.checked_by or "").strip(),
        "razorpay_link": None,
        "notes": data.notes,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
    }
    # Optional razorpay link
    if data.payment_method == "razorpay":
        link = await _create_razorpay_link(doc)
        if link:
            doc["razorpay_link"] = link
    await db.bills.insert_one(doc)
    # Upsert customer profile with history
    if data.customer_phone or data.customer_name:
        key = data.customer_phone or data.customer_name.lower()
        existing = await db.customers.find_one({"key": key})
        if existing:
            await db.customers.update_one({"key": key}, {"$set": {
                "name": data.customer_name,
                "phone": data.customer_phone,
                "email": data.customer_email or existing.get("email", ""),
                "last_visit": now_iso(),
            }, "$inc": {"visits": 1, "total_spent": doc["total"]}})
        else:
            await db.customers.insert_one({
                "id": new_id(),
                "key": key,
                "name": data.customer_name,
                "phone": data.customer_phone,
                "email": data.customer_email or "",
                "visits": 1,
                "total_spent": doc["total"],
                "first_visit": now_iso(),
                "last_visit": now_iso(),
            })
    doc.pop("_id", None)
    return doc

# ---------------- Customers ----------------
@api.get("/customers")
async def list_customers(user: dict = Depends(get_current_user)):
    return await db.customers.find({}, {"_id": 0}).sort("last_visit", -1).to_list(2000)

@api.get("/customers/{key}")
async def customer_detail(key: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one({"$or": [{"id": key}, {"key": key}]}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Not found")
    bills = await db.bills.find({"$or": [{"customer_phone": c["key"]}, {"customer_name": c["name"]}]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"customer": c, "bills": bills}

@api.patch("/bills/{bid}/status")
async def update_bill_status(bid: str, payload: dict, user: dict = Depends(get_current_user)):
    ps = payload.get("payment_status")
    if ps not in ("pending", "paid"):
        raise HTTPException(400, "Invalid status")
    await db.bills.update_one({"id": bid}, {"$set": {"payment_status": ps}})
    return await db.bills.find_one({"id": bid}, {"_id": 0})

@api.post("/bills/{bid}/send")
async def send_bill(bid: str, data: SendBillIn, user: dict = Depends(get_current_user)):
    bill = await db.bills.find_one({"id": bid}, {"_id": 0})
    if not bill:
        raise HTTPException(404, "Bill not found")
    settings = await _get_settings()
    msg = _format_bill_message(bill, settings)
    result = await _send_message(data.channel, bill.get("customer_phone", ""), bill.get("customer_email", ""), f"Your Bill {bill['bill_no']} from {settings.get('park_name','Funland')}", msg)
    return {"ok": True, "delivery": result}

# ---------------- Prebookings (public + internal) ----------------
def _prebook_no():
    return "BK-" + datetime.now(timezone.utc).strftime("%y%m%d") + "-" + uuid.uuid4().hex[:5].upper()

@api.get("/prebook/catalog")
async def prebook_catalog():
    """Public: list of active games + packages for the booking page."""
    games = await db.games.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(500)
    pkgs = await db.packages.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(500)
    s = await _get_settings()
    return {
        "park_name": s.get("park_name", "Funland Adventure Park"),
        "phone": s.get("phone", ""),
        "address": s.get("address", ""),
        "upi_qr_url": s.get("upi_qr_url", ""),
        "upi_id": s.get("upi_id", ""),
        "games": games,
        "packages": pkgs,
    }

@api.post("/prebook")
async def create_prebook(data: PrebookIn):
    """Public: create a prebooking. Returns booking id + payment info."""
    if not data.items:
        raise HTTPException(400, "At least one item required")
    total = round(sum(i.price * i.qty for i in data.items), 2)
    doc = {
        "id": new_id(),
        "booking_no": _prebook_no(),
        "customer_name": data.customer_name,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email or "",
        "booking_date": data.booking_date,
        "booking_time": data.booking_time or "",
        "pax": data.pax,
        "items": [i.model_dump() for i in data.items],
        "total": total,
        "notes": data.notes or "",
        "source": data.source,
        "status": "pending",
        "payment_status": "pending",
        "razorpay_link": None,
        "created_at": now_iso(),
    }
    # Try to create Razorpay link
    fake_bill = {
        "total": total,
        "bill_no": doc["booking_no"],
        "customer_name": data.customer_name,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email,
    }
    link = await _create_razorpay_link(fake_bill)
    if link:
        doc["razorpay_link"] = link
    await db.prebookings.insert_one(doc)
    doc.pop("_id", None)
    logger.info(f"Prebooking created: {doc['booking_no']} for {data.customer_name}")
    return doc

@api.get("/prebook/{bid}")
async def get_prebook_public(bid: str):
    """Public: get a prebooking by id OR booking_no — customer link opens this."""
    b = await db.prebookings.find_one({"$or": [{"id": bid}, {"booking_no": bid.upper()}]}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Booking not found")
    s = await _get_settings()
    b["_park"] = {"name": s.get("park_name"), "upi_qr_url": s.get("upi_qr_url"), "upi_id": s.get("upi_id"), "phone": s.get("phone"), "address": s.get("address")}
    return b

@api.get("/prebookings")
async def list_prebookings(user: dict = Depends(get_current_user)):
    return await db.prebookings.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)

@api.patch("/prebookings/{bid}/status")
async def update_prebook_status(bid: str, data: PrebookStatusUpdate, user: dict = Depends(get_current_user)):
    existing = await db.prebookings.find_one({"id": bid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Booking not found")
    # Once a prebooking has been converted to a bill, only admins can further edit it
    if existing.get("converted_bill_id") and user.get("role") != "admin":
        raise HTTPException(403, "Locked — this prebooking has already been billed. Only admin can edit it now.")
    upd = {"status": data.status}
    if data.status == "paid":
        upd["payment_status"] = "paid"
    if data.status == "cancelled":
        upd["payment_status"] = "cancelled"
    await db.prebookings.update_one({"id": bid}, {"$set": upd})
    return await db.prebookings.find_one({"id": bid}, {"_id": 0})

@api.post("/prebookings/{bid}/convert")
async def convert_prebook_to_bill(bid: str, user: dict = Depends(get_current_user)):
    """Convert a prebooking into an actual bill on customer arrival."""
    b = await db.prebookings.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Booking not found")
    items = [{"kind": it["kind"], "ref_id": it.get("ref_id"), "name": it["name"], "price": it["price"], "qty": it["qty"], "gst_percent": 0, "category": None, "hsn_code": ""} for it in b.get("items", [])]
    subtotal, disc_amount, gst_amount, total, gst_breakup = _compute_bill_totals(items, 0, 0, 0, False)
    bill_doc = {
        "id": new_id(),
        "bill_no": _bill_number(),
        "customer_name": b["customer_name"],
        "customer_phone": b["customer_phone"],
        "customer_email": b.get("customer_email", ""),
        "customer_gstin": "",
        "customer_state_code": "",
        "items": items,
        "discount": 0, "discount_percent": 0,
        "gst_percent": 0, "gst_amount": gst_amount, "gst_breakup": gst_breakup, "is_interstate": False,
        "subtotal": subtotal, "total": total,
        "payment_method": "razorpay" if b.get("razorpay_link") else "cash",
        "payment_status": "paid" if b.get("payment_status") == "paid" else "pending",
        "razorpay_link": b.get("razorpay_link"),
        "notes": f"From prebooking {b['booking_no']}",
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
        "prebooking_id": b["id"],
    }
    await db.bills.insert_one(bill_doc)
    await db.prebookings.update_one({"id": bid}, {"$set": {"status": "arrived", "converted_bill_id": bill_doc["id"]}})
    bill_doc.pop("_id", None)
    return bill_doc

@api.post("/prebook/{bid}/send")
async def send_prebook_link(bid: str, payload: dict, user: dict = Depends(get_current_user)):
    """Send prebooking link/QR to customer via WhatsApp/SMS/Email."""
    b = await db.prebookings.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Booking not found")
    channel = payload.get("channel", "whatsapp")
    frontend = os.environ.get("FRONTEND_URL", "").rstrip("/") or "https://game-package-tracker.preview.emergentagent.com"
    public_url = f"{frontend}/book/{b['booking_no']}"
    settings = await _get_settings()
    park = settings.get("park_name") or settings.get("firm_name") or "Funland Adventure Park"
    lines = [
        f"🎡 *{park}*",
        f"",
        f"Namaste {b['customer_name']}!",
        f"Aapki booking *{b['booking_no']}* confirm ho gayi ✅",
        f"Date: {b['booking_date']} {b.get('booking_time','')}",
        f"Amount: ₹{b['total']}",
        "",
        f"📋 View & Pay: {public_url}",
    ]
    if b.get("razorpay_link"):
        lines.append(f"💳 Pay online: {b['razorpay_link']}")
    if settings.get("upi_id"):
        lines.append(f"📲 UPI: {settings['upi_id']}")
    if settings.get("phone"):
        lines.append(f"📞 Contact: {settings['phone']}")
    lines.append("")
    lines.append("Thank you for choosing Funland! 🎉")
    msg = "\n".join(lines)
    res = await _send_message(channel, b.get("customer_phone", ""), b.get("customer_email", ""), f"{park} — Booking {b['booking_no']}", msg)
    return {"ok": True, "delivery": res, "public_url": public_url}

# ---------------- Attendance ----------------
@api.post("/attendance/checkin")
async def check_in(data: AttendanceCheckIn, user: dict = Depends(get_current_user)):
    today = date.today().isoformat()
    existing = await db.attendance.find_one({"user_id": user["id"], "date": today})
    if existing and existing.get("check_in"):
        raise HTTPException(400, "Already checked in today")
    doc = {
        "id": new_id(),
        "user_id": user["id"],
        "user_name": user["name"],
        "date": today,
        "check_in": now_iso(),
        "check_out": None,
        "notes": data.notes,
    }
    await db.attendance.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.post("/attendance/checkout")
async def check_out(user: dict = Depends(get_current_user)):
    today = date.today().isoformat()
    rec = await db.attendance.find_one({"user_id": user["id"], "date": today})
    if not rec:
        raise HTTPException(400, "No check-in found for today")
    if rec.get("check_out"):
        raise HTTPException(400, "Already checked out today")
    await db.attendance.update_one({"id": rec["id"]}, {"$set": {"check_out": now_iso()}})
    return await db.attendance.find_one({"id": rec["id"]}, {"_id": 0})

@api.get("/attendance/me")
async def my_attendance(user: dict = Depends(get_current_user)):
    return await db.attendance.find({"user_id": user["id"]}, {"_id": 0}).sort("date", -1).limit(60).to_list(60)

@api.get("/attendance/today")
async def today_attendance(user: dict = Depends(get_current_user)):
    today = date.today().isoformat()
    rec = await db.attendance.find_one({"user_id": user["id"], "date": today}, {"_id": 0})
    return rec

@api.get("/attendance/all")
async def all_attendance(_: dict = Depends(require_admin), days: int = 30):
    since = (date.today() - timedelta(days=days)).isoformat()
    return await db.attendance.find({"date": {"$gte": since}}, {"_id": 0}).sort("date", -1).to_list(2000)

# ---------------- Settings ----------------
async def _get_settings():
    s = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if not s:
        s = {
            "id": "global",
            "park_name": "Funland Adventure Park",
            "gst_rate": 0.0,
            "upi_qr_url": "",
            "upi_id": "",
            "phone": "",
            "address": "Indore, MP",
            "google_review_url": "",
            "google_reviews_shown": 0,
            "google_rating": 0.0,
            "firm_name": "",
            "firm_gstin": "",
            "firm_state_code": "23",
            "firm_pan": "",
            "firm_fssai": "",
            "invoice_prefix": "",
        }
        await db.settings.insert_one(s)
    # Backfill new fields for existing rows
    defaults = {
        "firm_name": s.get("park_name", ""),
        "firm_gstin": "",
        "firm_state_code": "23",
        "firm_pan": "",
        "firm_fssai": "",
        "invoice_prefix": "",
        "inquiry_webhook_secret": secrets.token_urlsafe(24),
        "meta_verify_token": secrets.token_urlsafe(16),
    }
    missing = {k: v for k, v in defaults.items() if k not in s or (k.endswith("_secret") or k.endswith("_token")) and not s.get(k)}
    if missing:
        await db.settings.update_one({"id": "global"}, {"$set": missing})
        s.update(missing)
    s.pop("_id", None)
    return s

async def _expand_shortlink(url: str) -> str:
    """Follow HTTP redirects for shortlinks and strip tracking params so QR encodes clean URL."""
    if not url or not url.startswith(("http://", "https://")):
        return url
    try:
        import httpx
        from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode
        async with httpx.AsyncClient(follow_redirects=True, timeout=6.0, headers={"User-Agent": "Mozilla/5.0"}) as c:
            r = await c.get(url)
            final = str(r.url)
            if not final or final == url:
                return url
            # Strip tracking / referral params from the expanded URL
            parsed = urlparse(final)
            keep_params = []
            drop_prefixes = ("utm_", "g_st", "g_ep", "coh", "entry", "skid")
            drop_keys = {"utm_source", "utm_medium", "utm_campaign", "g_st", "g_ep", "coh", "entry", "skid", "authuser", "hl"}
            for k, v in parse_qsl(parsed.query, keep_blank_values=False):
                if k.lower() in drop_keys or any(k.lower().startswith(p) for p in drop_prefixes):
                    continue
                keep_params.append((k, v))
            cleaned = urlunparse(parsed._replace(query=urlencode(keep_params)))
            logger.info(f"Expanded {url} -> {cleaned}")
            return cleaned
    except Exception as e:
        logger.warning(f"Shortlink expand failed for {url}: {e}")
    return url

@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    return await _get_settings()

@api.patch("/settings")
async def update_settings(data: SettingsIn, _: dict = Depends(require_admin)):
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    # Auto-expand Google review shortlink so QR + click always work
    if "google_review_url" in update and update["google_review_url"]:
        expanded = await _expand_shortlink(update["google_review_url"])
        update["google_review_url_original"] = update["google_review_url"]
        update["google_review_url"] = expanded
    if update:
        await db.settings.update_one({"id": "global"}, {"$set": update}, upsert=True)
    return await _get_settings()

# ---------------- Dashboard ----------------
@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    today = date.today().isoformat()
    since_7 = (date.today() - timedelta(days=7)).isoformat()

    bills_today = await db.bills.find({"created_at": {"$gte": today}}, {"_id": 0}).to_list(2000)
    revenue_today = sum(b.get("total", 0) for b in bills_today if b.get("payment_status") == "paid")
    footfall_today = len(bills_today)

    bills_week = await db.bills.find({"created_at": {"$gte": since_7}}, {"_id": 0, "total": 1, "created_at": 1, "payment_status": 1}).to_list(5000)

    _not_deleted = {"$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]}
    inquiries_new = await db.inquiries.count_documents({"status": "new", **_not_deleted})
    total_inquiries = await db.inquiries.count_documents(_not_deleted)
    pending_bills = await db.bills.count_documents({"payment_status": "pending"})
    pending_prebookings = await db.prebookings.count_documents({"status": {"$in": ["pending", "confirmed"]}})

    trend = {}
    for i in range(7):
        d = (date.today() - timedelta(days=i)).isoformat()
        trend[d] = 0
    for b in bills_week:
        d = b["created_at"][:10]
        if d in trend and b.get("payment_status") == "paid":
            trend[d] += b.get("total", 0)
    trend_list = [{"date": d, "revenue": round(v, 2)} for d, v in sorted(trend.items())]

    pipeline_games = await db.bills.find({}, {"_id": 0, "items": 1}).to_list(5000)
    game_counts = {}
    package_counts = {}
    total_games_played = 0
    total_packages_sold = 0
    games_revenue = 0.0
    packages_revenue = 0.0
    for b in pipeline_games:
        for it in b.get("items", []):
            qty = it.get("qty", 1)
            line_total = float(it.get("price", 0)) * qty
            if it.get("kind") == "game":
                game_counts[it["name"]] = game_counts.get(it["name"], 0) + qty
                total_games_played += qty
                games_revenue += line_total
            elif it.get("kind") == "package":
                package_counts[it["name"]] = package_counts.get(it["name"], 0) + qty
                total_packages_sold += qty
                packages_revenue += line_total
    top_games = sorted([{"name": k, "count": v} for k, v in game_counts.items()], key=lambda x: -x["count"])[:5]
    top_packages = sorted([{"name": k, "count": v} for k, v in package_counts.items()], key=lambda x: -x["count"])[:5]

    return {
        "revenue_today": round(revenue_today, 2),
        "footfall_today": footfall_today,
        "inquiries_new": inquiries_new,
        "total_inquiries": total_inquiries,
        "pending_bills": pending_bills,
        "pending_prebookings": pending_prebookings,
        "revenue_trend": trend_list,
        "top_games": top_games,
        "top_packages": top_packages,
        "total_games_played": total_games_played,
        "total_packages_sold": total_packages_sold,
        "games_revenue": round(games_revenue, 2),
        "packages_revenue": round(packages_revenue, 2),
    }

@api.get("/dashboard/analytics")
async def dashboard_analytics(
    from_date: str,
    to_date: str,
    granularity: Literal["day", "week", "month", "year"] = "day",
    user: dict = Depends(get_current_user),
):
    """Aggregated revenue + footfall between [from_date, to_date] inclusive, bucketed by granularity."""
    try:
        d_from = datetime.strptime(from_date, "%Y-%m-%d").date()
        d_to = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")
    if d_to < d_from:
        raise HTTPException(400, "to_date must be >= from_date")

    # inclusive end-of-day
    to_boundary = (d_to + timedelta(days=1)).isoformat()

    bills = await db.bills.find(
        {"created_at": {"$gte": d_from.isoformat(), "$lt": to_boundary}},
        {"_id": 0, "total": 1, "created_at": 1, "payment_status": 1, "items": 1, "customer_phone": 1},
    ).to_list(50000)

    def bucket_key(iso_str: str) -> str:
        d = datetime.fromisoformat(iso_str.replace("Z", "+00:00")).date()
        if granularity == "day":
            return d.isoformat()
        if granularity == "week":
            iso = d.isocalendar()
            return f"{iso[0]}-W{iso[1]:02d}"
        if granularity == "month":
            return f"{d.year}-{d.month:02d}"
        return str(d.year)

    def all_buckets():
        cur = d_from
        seen = []
        while cur <= d_to:
            k = bucket_key(cur.isoformat())
            if not seen or seen[-1] != k:
                seen.append(k)
            cur = cur + timedelta(days=1)
        return seen

    buckets_ordered = all_buckets()
    revenue_map = {k: 0.0 for k in buckets_ordered}
    footfall_map = {k: 0 for k in buckets_ordered}
    games_map = {k: 0 for k in buckets_ordered}
    packages_map = {k: 0 for k in buckets_ordered}
    total_revenue = 0.0
    total_footfall = 0
    total_paid = 0
    total_pending = 0
    game_counts = {}
    package_counts = {}
    total_games_played = 0
    total_packages_sold = 0
    games_revenue = 0.0
    packages_revenue = 0.0

    for b in bills:
        k = bucket_key(b["created_at"])
        if k in footfall_map:
            footfall_map[k] += 1
            total_footfall += 1
        if b.get("payment_status") == "paid":
            amt = b.get("total", 0)
            if k in revenue_map:
                revenue_map[k] += amt
            total_revenue += amt
            total_paid += 1
        else:
            total_pending += 1
        for it in b.get("items", []):
            qty = it.get("qty", 1)
            line_total = float(it.get("price", 0)) * qty
            if it.get("kind") == "game":
                game_counts[it["name"]] = game_counts.get(it["name"], 0) + qty
                total_games_played += qty
                games_revenue += line_total
                if k in games_map:
                    games_map[k] += qty
            elif it.get("kind") == "package":
                package_counts[it["name"]] = package_counts.get(it["name"], 0) + qty
                total_packages_sold += qty
                packages_revenue += line_total
                if k in packages_map:
                    packages_map[k] += qty

    trend = [{"date": k, "revenue": round(revenue_map[k], 2), "footfall": footfall_map[k], "games_played": games_map[k], "packages_sold": packages_map[k]} for k in buckets_ordered]
    top_games = sorted([{"name": k, "count": v} for k, v in game_counts.items()], key=lambda x: -x["count"])[:5]
    top_packages = sorted([{"name": k, "count": v} for k, v in package_counts.items()], key=lambda x: -x["count"])[:5]
    unique_customers = len({b.get("customer_phone", "") for b in bills if b.get("customer_phone")})

    return {
        "from": from_date,
        "to": to_date,
        "granularity": granularity,
        "total_revenue": round(total_revenue, 2),
        "total_footfall": total_footfall,
        "unique_customers": unique_customers,
        "bills_paid": total_paid,
        "bills_pending": total_pending,
        "average_bill": round(total_revenue / total_paid, 2) if total_paid else 0,
        "total_games_played": total_games_played,
        "total_packages_sold": total_packages_sold,
        "games_revenue": round(games_revenue, 2),
        "packages_revenue": round(packages_revenue, 2),
        "trend": trend,
        "top_games": top_games,
        "top_packages": top_packages,
    }

# ---------------- Marketing ----------------
@api.get("/campaigns")
async def list_campaigns(_: dict = Depends(require_admin)):
    return await db.campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api.post("/campaigns")
async def create_campaign(data: CampaignIn, admin: dict = Depends(require_admin)):
    # Determine audience phones
    phones: List[str] = []
    emails: List[str] = []
    if data.audience == "custom":
        phones = [p.strip() for p in data.custom_phones if p.strip()]
    elif data.audience == "all_customers":
        bills = await db.bills.find({}, {"_id": 0, "customer_phone": 1, "customer_email": 1}).to_list(5000)
        phones = list({b.get("customer_phone", "") for b in bills if b.get("customer_phone")})
        emails = list({b.get("customer_email", "") for b in bills if b.get("customer_email")})
    elif data.audience == "recent_customers":
        since = (date.today() - timedelta(days=30)).isoformat()
        bills = await db.bills.find({"created_at": {"$gte": since}}, {"_id": 0, "customer_phone": 1, "customer_email": 1}).to_list(5000)
        phones = list({b.get("customer_phone", "") for b in bills if b.get("customer_phone")})
        emails = list({b.get("customer_email", "") for b in bills if b.get("customer_email")})
    elif data.audience == "inquiries":
        inqs = await db.inquiries.find({"$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]}, {"_id": 0, "phone": 1, "email": 1}).to_list(5000)
        phones = list({i.get("phone", "") for i in inqs if i.get("phone")})
        emails = list({i.get("email", "") for i in inqs if i.get("email")})

    sent = 0
    failed = 0
    channel = data.channel
    if channel in ("instagram", "facebook"):
        # Save as content draft; social APIs not integrated
        sent = 0
    else:
        recipients = emails if channel == "email" else phones
        for r in recipients:
            res = await _send_message(channel, r if channel != "email" else "", r if channel == "email" else "", data.title, data.message)
            if res.get("ok"):
                sent += 1
            else:
                failed += 1

    doc = {
        "id": new_id(),
        "title": data.title,
        "channel": channel,
        "message": data.message,
        "image_url": data.image_url or "",
        "audience": data.audience,
        "target_count": len(emails) if channel == "email" else len(phones),
        "sent_count": sent,
        "failed_count": failed,
        "status": "draft" if channel in ("instagram", "facebook") else ("sent" if sent > 0 else "failed"),
        "created_by": admin["id"],
        "created_at": now_iso(),
    }
    await db.campaigns.insert_one(doc)
    doc.pop("_id", None)
    return doc

# ---------------- Integrations ----------------
def _integrations_status():
    return {
        "razorpay": bool(os.environ.get("RAZORPAY_KEY_ID") and os.environ.get("RAZORPAY_KEY_SECRET")),
        "twilio_sms": bool(os.environ.get("TWILIO_ACCOUNT_SID") and os.environ.get("TWILIO_AUTH_TOKEN") and os.environ.get("TWILIO_SMS_FROM")),
        "twilio_whatsapp": bool(os.environ.get("TWILIO_ACCOUNT_SID") and os.environ.get("TWILIO_AUTH_TOKEN") and os.environ.get("TWILIO_WHATSAPP_FROM")),
        "resend": bool(os.environ.get("RESEND_API_KEY")),
    }

@api.get("/integrations/status")
async def integrations_status(user: dict = Depends(get_current_user)):
    return _integrations_status()

async def _create_razorpay_link(bill: dict) -> Optional[str]:
    kid = os.environ.get("RAZORPAY_KEY_ID")
    ksec = os.environ.get("RAZORPAY_KEY_SECRET")
    if not (kid and ksec):
        return None
    try:
        import razorpay
        rc = razorpay.Client(auth=(kid, ksec))
        link = rc.payment_link.create({
            "amount": int(round(bill["total"] * 100)),
            "currency": "INR",
            "accept_partial": False,
            "description": f"Funland Bill {bill['bill_no']}",
            "customer": {
                "name": bill.get("customer_name", ""),
                "contact": bill.get("customer_phone", ""),
                "email": bill.get("customer_email", "") or None,
            },
            "notify": {"sms": bool(bill.get("customer_phone")), "email": bool(bill.get("customer_email"))},
            "reminder_enable": True,
        })
        return link.get("short_url")
    except Exception as e:
        logger.error(f"Razorpay error: {e}")
        return None

def _format_bill_message(bill: dict, settings: dict) -> str:
    park = settings.get("park_name", "Funland Adventure Park")
    lines = [
        f"*{park}*",
        f"Bill: {bill['bill_no']}",
        f"Customer: {bill['customer_name']}",
    ]
    if bill.get("customer_gstin"):
        lines.append(f"GSTIN: {bill['customer_gstin']}")
    lines.append("")
    for it in bill["items"]:
        gst = it.get("gst_percent") or 0
        tag = f" [{int(gst)}%]" if gst else ""
        lines.append(f"- {it['name']}{tag} x{it['qty']}  ₹{round(it['price']*it['qty'], 2)}")
    lines.append("")
    lines.append(f"Subtotal: ₹{bill['subtotal']}")
    if bill.get("discount"):
        lines.append(f"Discount: -₹{bill['discount']}")
    for br in (bill.get("gst_breakup") or []):
        rate = br.get("rate", 0)
        if bill.get("is_interstate"):
            lines.append(f"IGST @{int(rate)}%: ₹{br.get('igst', 0)}")
        else:
            lines.append(f"CGST @{rate/2:g}% + SGST @{rate/2:g}% ({int(rate)}% on ₹{br.get('taxable',0)}): ₹{br.get('total_tax', 0)}")
    if not bill.get("gst_breakup") and bill.get("gst_amount"):
        lines.append(f"GST ({bill.get('gst_percent', 0)}%): ₹{bill['gst_amount']}")
    lines.append(f"*Total: ₹{bill['total']}*")
    lines.append(f"Status: {bill['payment_status'].upper()}")
    if bill.get("razorpay_link"):
        lines.append(f"Pay online: {bill['razorpay_link']}")
    if settings.get("upi_id"):
        lines.append(f"UPI: {settings['upi_id']}")
    lines.append("")
    lines.append("Thank you for visiting!")
    if settings.get("google_review_url"):
        lines.append(f"⭐ Rate us on Google: {settings['google_review_url']}")
    return "\n".join(lines)

async def _send_message(channel: str, phone: str, email: str, subject: str, message: str) -> dict:
    """Try to send via configured provider; otherwise return simulated=True."""
    try:
        if channel in ("sms", "whatsapp"):
            sid = os.environ.get("TWILIO_ACCOUNT_SID")
            tok = os.environ.get("TWILIO_AUTH_TOKEN")
            frm = os.environ.get("TWILIO_SMS_FROM") if channel == "sms" else os.environ.get("TWILIO_WHATSAPP_FROM")
            if not (sid and tok and frm and phone):
                logger.info(f"[SIMULATED {channel}] to={phone} msg={message[:80]}")
                return {"ok": True, "simulated": True, "channel": channel}
            from twilio.rest import Client as TwilioClient
            client_t = TwilioClient(sid, tok)
            to = f"whatsapp:{phone}" if channel == "whatsapp" else phone
            from_ = f"whatsapp:{frm}" if channel == "whatsapp" else frm
            m = client_t.messages.create(body=message, from_=from_, to=to)
            return {"ok": True, "sid": m.sid, "channel": channel}
        if channel == "email":
            key = os.environ.get("RESEND_API_KEY")
            if not (key and email):
                logger.info(f"[SIMULATED email] to={email} sub={subject}")
                return {"ok": True, "simulated": True, "channel": "email"}
            import resend
            resend.api_key = key
            r = resend.Emails.send({
                "from": os.environ.get("RESEND_FROM_EMAIL", "Funland <onboarding@resend.dev>"),
                "to": [email],
                "subject": subject,
                "text": message,
            })
            return {"ok": True, "id": r.get("id"), "channel": "email"}
    except Exception as e:
        logger.error(f"send_message {channel} failed: {e}")
        return {"ok": False, "error": str(e), "channel": channel}
    return {"ok": False, "error": "unknown_channel"}

# ---------------- Bootstrap ----------------
async def seed_admin():
    email = os.environ.get("ADMIN_EMAIL", "admin@funland.in").lower()
    pw = os.environ.get("ADMIN_PASSWORD", "Funland@123")
    existing = await db.users.find_one({"email": email})
    if not existing:
        await db.users.insert_one({
            "id": new_id(),
            "email": email,
            "name": "Funland Manager",
            "phone": "",
            "role": "admin",
            "permissions": ALL_PERMS,
            "is_marketing_exec": False,
            "password_hash": hash_pw(pw),
            "created_at": now_iso(),
        })
        logger.info(f"Seeded admin: {email}")
    else:
        # Backfill missing fields for existing users
        upd = {}
        if "permissions" not in existing:
            upd["permissions"] = ALL_PERMS if existing.get("role") == "admin" else DEFAULT_EMP_PERMS
        if "is_marketing_exec" not in existing:
            upd["is_marketing_exec"] = False
        if upd:
            await db.users.update_one({"id": existing["id"]}, {"$set": upd})

async def ensure_indexes():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.games.create_index("id", unique=True)
    await db.packages.create_index("id", unique=True)
    await db.inquiries.create_index("id", unique=True)
    await db.bills.create_index("id", unique=True)
    await db.attendance.create_index([("user_id", 1), ("date", 1)])
    await db.prebookings.create_index("id", unique=True)
    await db.prebookings.create_index("booking_no", unique=True)

import asyncio

async def _startup_async():
    """Non-blocking startup - runs in background so /api endpoints answer fast even on cold start."""
    try:
        await ensure_indexes()
        await seed_admin()
        await _get_settings()
        logger.info("Funland CRM background init complete")
    except Exception as e:
        logger.error(f"Startup init failed (will retry on demand): {e}")

@app.on_event("startup")
async def startup():
    # Kick off init in background - do NOT block the HTTP server
    asyncio.create_task(_startup_async())
    logger.info("Funland CRM listening (background init in progress)")

@app.on_event("shutdown")
async def shutdown():
    client.close()

@api.get("/")
async def root():
    return {"service": "Funland CRM", "status": "ok"}

@api.get("/ping")
async def ping():
    """Ultra-lightweight health probe used by frontend heartbeat."""
    try:
        # tiny mongodb touch — verifies DB is reachable
        await db.command("ping")
        return {"ok": True, "ts": now_iso()}
    except Exception as e:
        raise HTTPException(503, f"db unavailable: {e}")

app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
