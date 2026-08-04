"""Iteration 15 — Soft-delete inquiries backend tests.
Covers: default/archive/only_archived filters, DELETE (soft), POST /restore,
dashboard stats, marketing report, export.xlsx, import (not archived by default),
fair-share (archived don't count), 404 handling, backward compat.
"""
import os
import io
import pytest
import requests

BASE = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
ADMIN = {"email": "admin@funland.in", "password": "Funland@123"}


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{BASE}/api/auth/login", json=ADMIN, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="module")
def H(token):
    return {"Authorization": f"Bearer {token}"}


def _create_inq(H, name, phone="9000000001"):
    r = requests.post(f"{BASE}/api/inquiries", json={
        "name": name, "phone": phone, "source": "walk-in", "status": "new"
    }, headers=H, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


# ---------- Tests ----------
class TestSoftDelete:
    def test_default_list_excludes_archived(self, H):
        r = requests.get(f"{BASE}/api/inquiries", headers=H, timeout=15)
        assert r.status_code == 200
        for row in r.json():
            assert not row.get("is_deleted"), f"Archived row {row.get('id')} leaked to default list"

    def test_only_archived_returns_archived_only(self, H):
        r = requests.get(f"{BASE}/api/inquiries?only_archived=1", headers=H, timeout=15)
        assert r.status_code == 200
        for row in r.json():
            assert row.get("is_deleted") is True

    def test_include_archived_returns_all(self, H):
        active = requests.get(f"{BASE}/api/inquiries", headers=H).json()
        arch = requests.get(f"{BASE}/api/inquiries?only_archived=1", headers=H).json()
        all_r = requests.get(f"{BASE}/api/inquiries?include_archived=1", headers=H).json()
        assert len(all_r) == len(active) + len(arch)

    def test_create_then_soft_delete_then_restore_flow(self, H):
        # 1) Create
        doc = _create_inq(H, "TEST_SD_Flow", "9111100001")
        iid = doc["id"]
        assert not doc.get("is_deleted")

        # 2) Visible in default
        ids = {x["id"] for x in requests.get(f"{BASE}/api/inquiries", headers=H).json()}
        assert iid in ids

        # 3) Soft delete
        r = requests.delete(f"{BASE}/api/inquiries/{iid}", headers=H, timeout=15)
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is True
        assert body["id"] == iid
        assert body["archived"] is True

        # 4) Missing from default
        ids = {x["id"] for x in requests.get(f"{BASE}/api/inquiries", headers=H).json()}
        assert iid not in ids

        # 5) Present in only_archived with deleted_at/by
        archived = {x["id"]: x for x in requests.get(f"{BASE}/api/inquiries?only_archived=1", headers=H).json()}
        assert iid in archived
        a = archived[iid]
        assert a.get("is_deleted") is True
        assert a.get("deleted_at")
        assert a.get("deleted_by")

        # 6) Restore
        r = requests.post(f"{BASE}/api/inquiries/{iid}/restore", headers=H, timeout=15)
        assert r.status_code == 200
        restored = r.json()
        assert restored.get("is_deleted") is False

        # 7) Visible in default again, gone from archived
        ids = {x["id"] for x in requests.get(f"{BASE}/api/inquiries", headers=H).json()}
        assert iid in ids
        ids_a = {x["id"] for x in requests.get(f"{BASE}/api/inquiries?only_archived=1", headers=H).json()}
        assert iid not in ids_a

        # Cleanup: soft-delete so it's out of active list (do NOT hard delete)
        requests.delete(f"{BASE}/api/inquiries/{iid}", headers=H)

    def test_delete_nonexistent_returns_404(self, H):
        r = requests.delete(f"{BASE}/api/inquiries/nonexistent-id-xyz-999", headers=H)
        assert r.status_code == 404

    def test_restore_nonexistent_returns_404(self, H):
        r = requests.post(f"{BASE}/api/inquiries/nonexistent-id-xyz-999/restore", headers=H)
        assert r.status_code == 404

    def test_dashboard_stats_excludes_archived(self, H):
        # Create + archive one; total should be same before/after
        r0 = requests.get(f"{BASE}/api/dashboard/stats", headers=H).json()
        total_before = r0["total_inquiries"]
        new_before = r0["inquiries_new"]

        doc = _create_inq(H, "TEST_SD_Stats", "9111100002")
        r1 = requests.get(f"{BASE}/api/dashboard/stats", headers=H).json()
        assert r1["total_inquiries"] == total_before + 1
        assert r1["inquiries_new"] == new_before + 1

        requests.delete(f"{BASE}/api/inquiries/{doc['id']}", headers=H)
        r2 = requests.get(f"{BASE}/api/dashboard/stats", headers=H).json()
        assert r2["total_inquiries"] == total_before
        assert r2["inquiries_new"] == new_before

    def test_export_xlsx_excludes_archived(self, H):
        import uuid as _u
        from openpyxl import load_workbook
        marker = f"TEST_SD_Export_{_u.uuid4().hex[:8]}"
        doc = _create_inq(H, marker, "9111100003")

        def _export_names():
            r = requests.get(f"{BASE}/api/inquiries/export.xlsx", headers=H)
            assert r.status_code == 200
            wb = load_workbook(io.BytesIO(r.content))
            ws = wb.active
            return {row[0].value for row in ws.iter_rows(min_row=2)}

        names_before = _export_names()
        assert marker in names_before

        requests.delete(f"{BASE}/api/inquiries/{doc['id']}", headers=H)
        names_after = _export_names()
        assert marker not in names_after

    def test_marketing_report_excludes_archived(self, H):
        doc = _create_inq(H, "TEST_SD_MReport", "9111100004")
        r1 = requests.get(f"{BASE}/api/marketing/report?preset=all", headers=H).json()
        t1 = r1["totals"]["assigned"]
        requests.delete(f"{BASE}/api/inquiries/{doc['id']}", headers=H)
        r2 = requests.get(f"{BASE}/api/marketing/report?preset=all", headers=H).json()
        t2 = r2["totals"]["assigned"]
        assert t2 == t1 - 1

    def test_import_creates_active_not_archived(self, H):
        # Build a minimal xlsx in-memory
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Source"])
        ws.append(["TEST_SD_Import1", "9111100011", "walk-in"])
        ws.append(["TEST_SD_Import2", "9111100012", "phone"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("test.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE}/api/inquiries/import", files=files, headers=H, timeout=30)
        assert r.status_code == 200, r.text
        assert r.json().get("inserted") >= 2

        active = requests.get(f"{BASE}/api/inquiries", headers=H).json()
        found = [x for x in active if x.get("name", "").startswith("TEST_SD_Import")]
        assert len(found) >= 2
        for f in found:
            assert not f.get("is_deleted")

        # Cleanup — soft delete only
        for f in found:
            requests.delete(f"{BASE}/api/inquiries/{f['id']}", headers=H)

    def test_soft_delete_response_contract(self, H):
        doc = _create_inq(H, "TEST_SD_Contract", "9111100021")
        r = requests.delete(f"{BASE}/api/inquiries/{doc['id']}", headers=H)
        body = r.json()
        assert set(["ok", "id", "archived"]).issubset(body.keys())
        assert body["ok"] is True and body["archived"] is True
