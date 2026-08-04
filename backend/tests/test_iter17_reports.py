"""Iteration 17: Reports, Expenses CRUD, business.xlsx, split print modes."""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://game-package-tracker.preview.emergentagent.com').rstrip('/')
API = f"{BASE_URL}/api"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": "admin@funland.in", "password": "Funland@123"})
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="module")
def h(token):
    return {"Authorization": f"Bearer {token}"}


# ---- Sales report ----
def test_sales_report_month(h):
    r = requests.get(f"{API}/reports/sales?preset=month", headers=h)
    assert r.status_code == 200
    d = r.json()
    for k in ["from", "to", "label", "total_revenue", "paid_bills", "pending_bills", "avg_bill_value", "category_totals", "daily", "top_items"]:
        assert k in d, f"missing {k}"
    assert isinstance(d["daily"], list)
    assert isinstance(d["top_items"], list)
    assert isinstance(d["category_totals"], dict)


def test_sales_report_custom_range(h):
    r = requests.get(f"{API}/reports/sales?from=2025-01-01&to=2026-12-31", headers=h)
    assert r.status_code == 200
    assert "total_revenue" in r.json()


# ---- GSTR-3B ----
def test_gstr3b_report(h):
    r = requests.get(f"{API}/reports/gstr3b?preset=month", headers=h)
    assert r.status_code == 200
    d = r.json()
    for k in ["invoice_count", "total_taxable", "total_cgst", "total_sgst", "total_igst", "total_tax", "rate_wise"]:
        assert k in d
    assert isinstance(d["rate_wise"], list)
    # rate_wise sorted by rate
    if len(d["rate_wise"]) > 1:
        rates = [r["rate"] for r in d["rate_wise"]]
        assert rates == sorted(rates)


# ---- Payment mode ----
def test_payment_mode(h):
    r = requests.get(f"{API}/reports/payment-mode?preset=month", headers=h)
    assert r.status_code == 200
    d = r.json()
    for k in ["total_paid", "total_pending", "modes"]:
        assert k in d
    for m in d["modes"]:
        for k in ["method", "paid_count", "paid_amount", "pending_count", "pending_amount"]:
            assert k in m


# ---- Expense CRUD ----
def test_expense_crud(h):
    payload = {"date": "2026-01-15", "category": "marketing", "amount": 500.0, "vendor": "TEST_ITER17_Vendor",
               "payment_method": "cash", "description": "TEST_ITER17 expense"}
    r = requests.post(f"{API}/expenses", json=payload, headers=h)
    assert r.status_code == 200, r.text
    eid = r.json()["id"]
    assert eid

    # List includes it
    r2 = requests.get(f"{API}/expenses", headers=h)
    assert r2.status_code == 200
    ids = [e["id"] for e in r2.json()]
    assert eid in ids

    # Patch
    payload["amount"] = 750.0
    r3 = requests.patch(f"{API}/expenses/{eid}", json=payload, headers=h)
    assert r3.status_code == 200
    assert r3.json()["amount"] == 750.0

    # Delete
    r4 = requests.delete(f"{API}/expenses/{eid}", headers=h)
    assert r4.status_code == 200


def test_expense_categories_accepted(h):
    ids = []
    for cat in ["rent", "salary", "utility", "food", "maintenance", "marketing", "travel", "other"]:
        r = requests.post(f"{API}/expenses", json={"date": "2026-01-15", "category": cat, "amount": 10.0,
                                                    "description": f"TEST_ITER17_{cat}"}, headers=h)
        assert r.status_code == 200, f"{cat}: {r.text}"
        ids.append(r.json()["id"])
    # cleanup
    for eid in ids:
        requests.delete(f"{API}/expenses/{eid}", headers=h)


def test_expense_report(h):
    # seed
    r = requests.post(f"{API}/expenses", json={"date": "2026-01-15", "category": "rent", "amount": 1000.0,
                                                "description": "TEST_ITER17_rpt"}, headers=h)
    eid = r.json()["id"]
    try:
        rep = requests.get(f"{API}/reports/expenses?preset=month", headers=h)
        assert rep.status_code == 200
        d = rep.json()
        for k in ["total", "count", "by_category", "by_month", "expenses"]:
            assert k in d
    finally:
        requests.delete(f"{API}/expenses/{eid}", headers=h)


# ---- Business XLSX ----
def test_business_xlsx(h):
    r = requests.get(f"{API}/reports/business.xlsx?preset=month", headers=h)
    assert r.status_code == 200
    assert r.content[:2] == b"PK", "Not a valid xlsx (missing PK magic)"
    assert len(r.content) > 5000
    # check sheets
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"Sales", "GSTR-3B", "Payment Modes", "Expenses"}


# ---- Auth guard ----
def test_expenses_requires_auth():
    r = requests.get(f"{API}/expenses")
    assert r.status_code == 401


def test_reports_requires_auth():
    r = requests.get(f"{API}/reports/sales?preset=month")
    assert r.status_code == 401
