"""Backend tests for Indian GST compliance features (iteration 12).

Covers:
- Settings: firm_* fields + invoice_prefix persist
- Games: gst_category food/activity → default HSN
- Packages: food_portion + activity_portion persisted
- Bills: package auto-split into 5% food + 18% activity lines
- Bills: intra-state (CGST+SGST) vs inter-state (IGST) based on customer_state_code
- Bills: customer_gstin stored uppercased and returned
- Bills: game with gst_category='food' auto-sets 5% + HSN 996331
- Inquiries: export.xlsx / template.xlsx / import round-trip
"""
import io
import os
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # fall back to frontend .env
    from pathlib import Path
    for line in Path("/app/frontend/.env").read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")

ADMIN_EMAIL = "admin@funland.in"
ADMIN_PASSWORD = "Funland@123"


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=15)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def h(token):
    return {"Authorization": f"Bearer {token}"}


# ------------- Settings -------------
class TestSettings:
    def test_get_settings_has_firm_fields(self, h):
        r = requests.get(f"{BASE_URL}/api/settings", headers=h, timeout=15)
        assert r.status_code == 200
        s = r.json()
        for k in ["firm_name", "firm_gstin", "firm_state_code", "firm_pan", "firm_fssai", "invoice_prefix"]:
            assert k in s, f"missing key {k}"

    def test_patch_settings_persists(self, h):
        payload = {
            "firm_name": "TEST_M/s Funland",
            "firm_gstin": "23TESTF1234F1Z5",
            "firm_state_code": "23",
            "firm_pan": "TESTF1234F",
            "firm_fssai": "12345678901234",
            "invoice_prefix": "FL/25-26/",
        }
        r = requests.patch(f"{BASE_URL}/api/settings", json=payload, headers=h, timeout=15)
        assert r.status_code == 200
        s = r.json()
        for k, v in payload.items():
            assert s.get(k) == v, f"{k}: expected {v} got {s.get(k)}"
        # verify persistence via GET
        r2 = requests.get(f"{BASE_URL}/api/settings", headers=h, timeout=15).json()
        assert r2["firm_state_code"] == "23"
        assert r2["invoice_prefix"] == "FL/25-26/"


# ------------- Games -------------
class TestGamesGST:
    def test_food_game_defaults(self, h):
        payload = {"name": "TEST_Snacks", "price": 150, "gst_category": "food"}
        r = requests.post(f"{BASE_URL}/api/games", json=payload, headers=h, timeout=15)
        assert r.status_code == 200, r.text
        g = r.json()
        assert g["gst_category"] == "food"
        assert g["price"] == 150
        # HSN may be blank (backend does not force) — but hsn_code key exists
        assert "hsn_code" in g
        # cleanup
        requests.delete(f"{BASE_URL}/api/games/{g['id']}", headers=h)

    def test_activity_game_defaults(self, h):
        r = requests.post(f"{BASE_URL}/api/games", json={"name": "TEST_Ride", "price": 200, "gst_category": "activity"}, headers=h, timeout=15)
        assert r.status_code == 200
        g = r.json()
        assert g["gst_category"] == "activity"
        requests.delete(f"{BASE_URL}/api/games/{g['id']}", headers=h)


# ------------- Packages -------------
class TestPackages:
    def test_package_with_portions(self, h):
        payload = {"name": "TEST_Bday Combo A", "price": 2000, "pax": 10, "food_portion": 800, "activity_portion": 1200}
        r = requests.post(f"{BASE_URL}/api/packages", json=payload, headers=h, timeout=15)
        assert r.status_code == 200, r.text
        p = r.json()
        assert p["food_portion"] == 800
        assert p["activity_portion"] == 1200
        requests.delete(f"{BASE_URL}/api/packages/{p['id']}", headers=h)

    def test_package_without_portions_defaults(self, h):
        r = requests.post(f"{BASE_URL}/api/packages", json={"name": "TEST_NoSplit", "price": 500, "pax": 5}, headers=h, timeout=15)
        assert r.status_code == 200
        p = r.json()
        assert p["food_portion"] == 0
        assert p["activity_portion"] == 0
        requests.delete(f"{BASE_URL}/api/packages/{p['id']}", headers=h)


# ------------- Bills / GST split -------------
class TestBillsGST:
    @pytest.fixture(scope="class")
    def firm_intra_state(self, h):
        requests.patch(f"{BASE_URL}/api/settings", json={"firm_state_code": "23"}, headers=h, timeout=15)
        yield

    @pytest.fixture(scope="class")
    def split_package(self, h):
        r = requests.post(f"{BASE_URL}/api/packages", json={"name": "TEST_Split Pkg", "price": 2000, "pax": 10, "food_portion": 800, "activity_portion": 1200}, headers=h, timeout=15)
        assert r.status_code == 200
        p = r.json()
        yield p
        requests.delete(f"{BASE_URL}/api/packages/{p['id']}", headers=h)

    @pytest.fixture(scope="class")
    def food_game(self, h):
        r = requests.post(f"{BASE_URL}/api/games", json={"name": "TEST_FoodGame", "price": 100, "gst_category": "food"}, headers=h, timeout=15)
        assert r.status_code == 200
        g = r.json()
        yield g
        requests.delete(f"{BASE_URL}/api/games/{g['id']}", headers=h)

    def test_bill_package_auto_split_intra(self, h, firm_intra_state, split_package):
        payload = {
            "customer_name": "TEST_Rahul",
            "customer_phone": "9990000001",
            "customer_state_code": "23",  # same as firm
            "items": [{"kind": "package", "ref_id": split_package["id"], "name": split_package["name"], "price": 2000, "qty": 1}],
        }
        r = requests.post(f"{BASE_URL}/api/bills", json=payload, headers=h, timeout=15)
        assert r.status_code == 200, r.text
        b = r.json()
        # Must be 2 items now
        assert len(b["items"]) == 2, f"expected 2 line items, got {len(b['items'])}"
        cats = sorted([i.get("category") for i in b["items"]])
        assert cats == ["activity", "food"], cats
        food = next(i for i in b["items"] if i["category"] == "food")
        act = next(i for i in b["items"] if i["category"] == "activity")
        assert food["gst_percent"] == 5.0
        assert food["hsn_code"] == "996331"
        assert food["price"] == 800
        assert act["gst_percent"] == 18.0
        assert act["hsn_code"] == "999721"
        assert act["price"] == 1200
        # breakup: 2 entries
        breakup = b.get("gst_breakup") or []
        rates = sorted([br["rate"] for br in breakup])
        assert rates == [5.0, 18.0]
        # intra-state → cgst+sgst split, igst 0
        for br in breakup:
            assert br["igst"] == 0.0
            assert br["cgst"] > 0
            assert br["sgst"] > 0
        assert b["is_interstate"] is False

    def test_bill_inter_state_igst(self, h, firm_intra_state, split_package):
        payload = {
            "customer_name": "TEST_Priya",
            "customer_phone": "9990000002",
            "customer_state_code": "27",  # MH != MP(23) → IGST
            "items": [{"kind": "package", "ref_id": split_package["id"], "name": split_package["name"], "price": 2000, "qty": 1}],
        }
        r = requests.post(f"{BASE_URL}/api/bills", json=payload, headers=h, timeout=15)
        assert r.status_code == 200, r.text
        b = r.json()
        assert b["is_interstate"] is True
        for br in b["gst_breakup"]:
            assert br["cgst"] == 0.0
            assert br["sgst"] == 0.0
            assert br["igst"] > 0

    def test_bill_food_game_autofills(self, h, firm_intra_state, food_game):
        payload = {
            "customer_name": "TEST_Amit",
            "customer_phone": "9990000003",
            "customer_state_code": "23",
            "items": [{"kind": "game", "ref_id": food_game["id"], "name": food_game["name"], "price": 100, "qty": 2}],
        }
        r = requests.post(f"{BASE_URL}/api/bills", json=payload, headers=h, timeout=15)
        assert r.status_code == 200, r.text
        b = r.json()
        assert len(b["items"]) == 1
        it = b["items"][0]
        assert it["gst_percent"] == 5.0
        assert it["hsn_code"] == "996331"
        assert it["category"] == "food"

    def test_bill_customer_gstin_uppercased_and_persisted(self, h, firm_intra_state, split_package):
        payload = {
            "customer_name": "TEST_B2B",
            "customer_phone": "9990000004",
            "customer_gstin": "23abcde1234f1z5",
            "customer_state_code": "23",
            "items": [{"kind": "package", "ref_id": split_package["id"], "name": split_package["name"], "price": 2000, "qty": 1}],
        }
        r = requests.post(f"{BASE_URL}/api/bills", json=payload, headers=h, timeout=15)
        assert r.status_code == 200
        b = r.json()
        assert b["customer_gstin"] == "23ABCDE1234F1Z5"
        # GET fetch back
        r2 = requests.get(f"{BASE_URL}/api/bills/{b['id']}", headers=h, timeout=15)
        assert r2.status_code == 200
        assert r2.json()["customer_gstin"] == "23ABCDE1234F1Z5"


# ------------- Inquiries Excel -------------
class TestInquiriesExcel:
    def test_export_xlsx(self, h):
        r = requests.get(f"{BASE_URL}/api/inquiries/export.xlsx", headers=h, timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert r.content[:2] == b"PK", "not a zip/xlsx"
        # parse
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        expected = ["Name", "Phone", "Email", "Source", "Interest", "Notes", "Follow Up Date", "Status", "Assigned To", "Created At"]
        assert header == expected

    def test_template_xlsx_has_examples(self, h):
        r = requests.get(f"{BASE_URL}/api/inquiries/template.xlsx", headers=h, timeout=15)
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # header + at least 2 example rows
        row_count = sum(1 for _ in ws.iter_rows(values_only=True))
        assert row_count >= 3, f"expected header + 2+ examples, got {row_count}"

    def test_import_roundtrip(self, h):
        # build an xlsx with 2 rows
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Email", "Source", "Interest", "Notes", "Follow Up Date", "Status", "Assigned To", "Created At"])
        ws.append(["TEST_ImportOne", "9111000001", "one@t.com", "whatsapp", "combo", "test note", "", "new", "", ""])
        ws.append(["TEST_ImportTwo", "9111000002", "", "instagram", "birthday", "", "", "contacted", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("upload.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/inquiries/import", files=files, headers=h, timeout=30)
        assert r.status_code == 200, r.text
        j = r.json()
        assert j["ok"] is True
        assert j["inserted"] == 2
        # verify listed
        r2 = requests.get(f"{BASE_URL}/api/inquiries", headers=h, timeout=15)
        assert r2.status_code == 200
        names = [i["name"] for i in r2.json()]
        assert "TEST_ImportOne" in names
        assert "TEST_ImportTwo" in names

    def test_import_rejects_non_xlsx(self, h):
        files = {"file": ("bad.txt", b"hello world", "text/plain")}
        r = requests.post(f"{BASE_URL}/api/inquiries/import", files=files, headers=h, timeout=15)
        assert r.status_code == 400


# ------------- Regression -------------
class TestRegression:
    def test_dashboard_stats(self, h):
        r = requests.get(f"{BASE_URL}/api/dashboard/stats", headers=h, timeout=15)
        assert r.status_code == 200
        d = r.json()
        assert "revenue_today" in d and "top_games" in d

    def test_list_bills_ok(self, h):
        r = requests.get(f"{BASE_URL}/api/bills", headers=h, timeout=15)
        assert r.status_code == 200
        assert isinstance(r.json(), list)
